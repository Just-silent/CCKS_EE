# -*- coding: utf-8 -*-
# @Author   : Just-silent
# @time     : 2020/8/8 20:34

import random
from openpyxl import load_workbook, Workbook

def data_argmentation():
    # 需要清洗
    vocab_train_path = './task2_vocab.txt'
    vocab_val_path = './task2_vocab.val.txt'
    train_data_path = './task2_train_reformat.xlsx'
    vocab_list = []
    for path in [vocab_train_path, vocab_val_path]:
        with open(path, 'r', encoding='utf-8') as vocab:
            for x in vocab.readlines():
                vocab_list.append(x[:-1].replace('_x0004_', '').replace(' ', ''))
    wb = load_workbook(train_data_path)
    ws = wb['sheet1']
    max_row = ws.max_row
    for i in range(max_row-1):
        line = i+1
        if ws.cell(line, 2).value is not None:
            places = ws.cell(line, 2).value.split(',')
            vocab_list.extend(places)
        if ws.cell(line, 4).value is not None:
            places = ws.cell(line, 4).value.split(',')
            vocab_list.extend(places)
    vocab_list = list(set(vocab_list))
    wb1 = load_workbook('./sub_cut_train.xlsx')
    ws1 = wb1['sheet1']
    wb2 = load_workbook('./sub_cut_dev.xlsx')
    ws2 =wb2['sheet1']
    max1_row = ws1.max_row
    max2_row = ws2.max_row
    names = ['原文', '原发部位', '病灶大小', '转移部位']
    wb1_new = Workbook()
    ws1_new = wb1_new.create_sheet('sheet1')
    wb1_new.remove(wb1_new['Sheet'])
    wb2_new = Workbook()
    ws2_new = wb2_new.create_sheet('sheet1')
    wb2_new.remove(wb2_new['Sheet'])
    for i in range(len(names)):
        ws1_new.cell(1, i + 1, names[i])
        ws2_new.cell(1, i + 1, names[i])
    line_1 = 2
    line_2 = 2
    rate = 3
    for i in range(max1_row-1):
        line = i+2
        sentence = ws1.cell(line, 1).value
        size = ws1.cell(line, 3).value
        if ws1.cell(line, 2).value is None and ws1.cell(line, 4).value is None:
            ws1_new.cell(line_1, 1, sentence)
            ws1_new.cell(line_1, 2, None)
            ws1_new.cell(line_1, 3, size)
            ws1_new.cell(line_1, 3, None)
            line_1+=1
        if ws1.cell(line, 2).value is not None and ws1.cell(line, 4).value is None:
            place = ws1.cell(line, 2).value
            ws1_new.cell(line_1, 1, sentence)
            ws1_new.cell(line_1, 2, place)
            ws1_new.cell(line_1, 3, size)
            line_1+=1
            places_random = random.sample(vocab_list, rate)
            sentences  = []
            for i in range(rate):
                sentences.append(sentence.replace(place, places_random[i]))
            for i in range(rate):
                ws1_new.cell(line_1, 1, sentences[i])
                ws1_new.cell(line_1, 2, places_random[i])
                ws1_new.cell(line_1, 3, size)
                line_1+=1
        if ws1.cell(line, 2).value is None and ws1.cell(line, 4).value is not None:
            place = ws1.cell(line, 4).value
            ws1_new.cell(line_1, 1, sentence)
            ws1_new.cell(line_1, 4, place)
            ws1_new.cell(line_1, 3, size)
            line_1+=1
            places_random = random.sample(vocab_list, rate)
            sentences = []
            for i in range(rate):
                sentences.append(sentence.replace(place, places_random[i]))
            for i in range(rate):
                ws1_new.cell(line_1, 1, sentences[i])
                ws1_new.cell(line_1, 4, places_random[i])
                ws1_new.cell(line_1, 3, size)
                line_1 += 1
        if ws1.cell(line, 2).value is not None and ws1.cell(line, 4).value is not None:
            place1 = ws1.cell(line, 2).value
            place2 = ws1.cell(line, 4).value
            ws1_new.cell(line_1, 1, sentence)
            ws1_new.cell(line_1, 2, place1)
            ws1_new.cell(line_1, 3, size)
            ws1_new.cell(line_1, 4, place2)
            line_1+=1
            places_random1 = random.sample(vocab_list, rate)
            places_random2 = random.sample(vocab_list, rate)
            sentences = []
            for i in range(rate):
                sentences.append(sentence.replace(place1, places_random1[i]).replace(place2, places_random2[i]))
            for i in range(rate):
                ws1_new.cell(line_1, 1, sentences[i])
                ws1_new.cell(line_1, 2, places_random1[i])
                ws1_new.cell(line_1, 3, size)
                ws1_new.cell(line_1, 4, places_random2[i])
                line_1 += 1
    print('max_train_row:{} line_1:{}'.format(max1_row, line_1))
    for i in range(max2_row-1):
        line = i+2
        sentence = ws2.cell(line, 1).value
        size = ws2.cell(line, 3).value
        if ws2.cell(line, 2).value is None and ws2.cell(line, 4).value is None:
            ws2_new.cell(line_2, 1, sentence)
            ws2_new.cell(line_2, 2, None)
            ws2_new.cell(line_2, 3, size)
            ws2_new.cell(line_2, 3, None)
            line_1+=1
        if ws2.cell(line, 2).value is not None and ws2.cell(line, 4).value is None:
            place = ws2.cell(line, 2).value
            ws2_new.cell(line_2, 1, sentence)
            ws2_new.cell(line_2, 2, place)
            ws2_new.cell(line_2, 3, size)
            line_2+=1
            places_random = random.sample(vocab_list, rate)
            sentences  = []
            for i in range(rate):
                sentences.append(sentence.replace(place, places_random[i]))
            for i in range(rate):
                ws2_new.cell(line_2, 1, sentences[i])
                ws2_new.cell(line_2, 2, places_random[i])
                ws2_new.cell(line_2, 3, size)
                line_2+=1
        if ws2.cell(line, 2).value is None and ws2.cell(line, 4).value is not None:
            place = ws2.cell(line, 4).value
            ws2_new.cell(line_2, 1, sentence)
            ws2_new.cell(line_2, 4, place)
            ws2_new.cell(line_2, 3, size)
            line_2+=1
            places_random = random.sample(vocab_list, rate)
            sentences = []
            for i in range(rate):
                sentences.append(sentence.replace(place, places_random[i]))
            for i in range(rate):
                ws2_new.cell(line_2, 1, sentences[i])
                ws2_new.cell(line_2, 4, places_random[i])
                ws2_new.cell(line_2, 3, size)
                line_2 += 1
        if ws2.cell(line, 2).value is not None and ws2.cell(line, 4).value is not None:
            place1 = ws2.cell(line, 2).value
            place2 = ws2.cell(line, 4).value
            ws1_new.cell(line_2, 1, sentence)
            ws1_new.cell(line_2, 2, place1)
            ws1_new.cell(line_2, 3, size)
            ws1_new.cell(line_2, 4, place2)
            line_2+=1
            places_random1 = random.sample(vocab_list, rate)
            places_random2 = random.sample(vocab_list, rate)
            sentences = []
            for i in range(rate):
                sentences.append(sentence.replace(place1, places_random1[i]).replace(place2, places_random2[i]))
            for i in range(rate):
                ws2_new.cell(line_2, 1, sentences[i])
                ws2_new.cell(line_2, 2, places_random1[i])
                ws2_new.cell(line_2, 3, size)
                ws2_new.cell(line_2, 4, places_random2[i])
                line_2 += 1
    print('max_dev_row:{} line_2:{}'.format(max2_row, line_2))
    wb1_new.save('./sub_cut_train_new.xlsx')
    wb2_new.save('./sub_cut_dev_new.xlsx')

if __name__ == '__main__':
    data_argmentation()
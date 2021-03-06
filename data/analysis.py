# -*- coding: utf-8 -*-
# @Author   : Just-silent
# @time     : 2020/5/2 9:00

import json
from tool import tool
from tqdm import tqdm
import numpy as np
from config import config
import matplotlib.pyplot as plt
from openpyxl import load_workbook


# 句子长度、占比
def get_lengths(path='./task2_train_reformat.xlsx'):
    wb = load_workbook(path)
    ws = wb['sheet1']
    max_row = ws.max_row
    sentences = []
    for i in range(max_row-2):
        # 除去第一个‘。’之前的字符
        # sentences.append(len(ws.cell(i+2, 1).value)-len(ws.cell(i+2, 1).value.split('。')[0]))
        sentences.append(len(ws.cell(i+2, 1).value))
    sen_in_inter = []
    interval = 20
    min_len = min(sentences)
    max_len = max(sentences)
    start = min_len // interval *interval
    interval_num = max_len//interval - min_len//interval + 1
    for i in range(interval_num):
        sen_in_inter.append(0)
    for i in range(len(sentences)):
        sen_in_inter[(sentences[i] - start) // interval]+=1
    print('区间个数：{}   每个区间数目：{}'.format(interval_num, sen_in_inter))
    print("区间为：",end='')
    x_names = []
    for i in range(interval_num):
        x_names.append('{}-{}'.format(start+interval*i, start+interval*(i+1)-1))
    print(x_names)

    # 这两行代码解决 plt 中文显示的问题
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False
    bar_width = 0.5  # 条形宽度
    # index_support = np.arange(len(x_name)) * 10  # support条形图的横坐标
    # 使用两次 bar 函数画出两组条形图
    rects = plt.bar(x_names, height=sen_in_inter, width=bar_width, color='b', label='句子数')
    # 显示对应柱状图的数值
    for rect in rects:
        height = rect.get_height()
        plt.text(rect.get_x() + rect.get_width() / 2, height, str(height), ha='center', va='bottom')
    plt.legend()  # 显示图例
    x = []
    for i in range(interval_num):
        x.append(i)
    plt.xticks(x, x_names, rotation=0)
    plt.ylabel('该范围句子个数')  # 纵坐标轴标题
    plt.title('句子长度分布')  # 图形标题
    plt.show()

def tagging_notagging():
    wb = load_workbook('./task2_train_reformat.xlsx')
    ws = wb['sheet1']
    max_row = ws.max_row
    tagging_num = 0
    notagging_num = 0
    for i in range(max_row-1):
        line = i+2
        if ws.cell(line,2).value is None and ws.cell(line,2).value is None and ws.cell(line,2).value is None:
            notagging_num+=1
        else:
            tagging_num+=1
    print('标注：{}    未标注：{}'.format(tagging_num, notagging_num))

def tags_proportion():
    wb = load_workbook('./task2_train_reformat.xlsx')
    ws = wb['sheet1']
    max_row = ws.max_row
    tags_num = [0,0,0]
    max = 0
    max_tags = []
    kind = 0
    sentence_proportion_tag = np.array(np.zeros((3,19)))
    for i in range(max_row - 1):
        line = i + 2
        for j in range(3):
            if ws.cell(line, j+2).value is not None:
                tag_len = len(ws.cell(line, j + 2).value.split(','))
                sentence_proportion_tag[j][tag_len] += 1
                if tag_len>max:
                    max = tag_len
                    max_tags = ws.cell(line, j + 2).value.split(',')
                    kind = j
                tags_num[j] += tag_len
            else:
                sentence_proportion_tag[j][0] += 1
    tags_proportion = []
    for i in range(3):
        tags_proportion.append(round(tags_num[i]/sum(tags_num), 2))
    print('原发部位占比：{}%  病灶大小占比：{}%  转移部位占比：{}%'.format(tags_proportion[0]*100, tags_proportion[1]*100, tags_proportion[2]*100))
    print('属性对多标签为：{}   有哪些：{}   类别：{}'.format(max, max_tags, ['原发部位 ', '病灶大小', '转移部位'][kind]))
    print(sentence_proportion_tag)

# 每个标签在句子中出现的平均次数
def analysis_tag_in_sentence(path='task2_train_reformat.xlsx'):
    wb = load_workbook(path)
    ws = wb['sheet1']
    max_row = ws.max_row
    tags = 0
    origin_tags = 0
    origin_num_in_sentence = 0
    origin_one_in_sentence = 0
    transfs_tags = 0
    transfs_num_in_sentence = 0
    transfs_one_in_sentence = 0
    size_tags = 0
    size_num_in_sentence = 0
    size_one_in_sentence = 0
    for i in range(max_row-1):
        row = i+2
        text = ws.cell(row, 1).value
        if ws.cell(row, 2).value is not None:
            origins = ws.cell(row, 2).value.split(',')
            origin_tags = origin_tags + len(origins)
            for name in origins:
                origin_num_in_sentence += len(tool.find_all_index(text, name))
                if len(tool.find_all_index(text, name)) ==1:
                    origin_one_in_sentence += 1
        if ws.cell(row, 4).value is not None:
            transfs = ws.cell(row, 4).value.split(',')
            transfs_tags = transfs_tags + len(transfs)
            for name in transfs:
                transfs_num_in_sentence += len(tool.find_all_index(text,name))
                if len(tool.find_all_index(text, name)) == 1:
                    transfs_one_in_sentence += 1
        if ws.cell(row, 3).value is not None:
            sizes = ws.cell(row, 3).value.split(',')
            size_tags = size_tags + len(sizes)
            for name in sizes:
                size_num_in_sentence += len(tool.find_all_index(text,name))
                if len(tool.find_all_index(text, name)) == 1:
                    size_one_in_sentence += 1
    print('原发部位部位每个标签句子中出现平均次数：', origin_num_in_sentence/origin_tags)
    print('原发部位每个标签句子中出现一次占比：', origin_one_in_sentence/origin_tags)
    print('转移部位每个标签句子中出现平均次数：', transfs_num_in_sentence/transfs_tags)
    print('转移部位每个标签句子中出现一次占比：', transfs_one_in_sentence/transfs_tags)
    print('部位大小每个标签句子中出现平均次数：', size_num_in_sentence/size_tags)
    print('部位大小每个标签句子中出现一次占比：', size_one_in_sentence/size_tags)
    print('三个属性每个标签句子中出现平均次数：', (origin_num_in_sentence + transfs_num_in_sentence + size_num_in_sentence) / (origin_tags + transfs_tags + size_tags))
    print('三个属性每个标签句子中出现一次占比：', (origin_one_in_sentence + transfs_one_in_sentence+ size_one_in_sentence) / (origin_tags + transfs_tags + size_tags))

def origin_proportion_i_sentence(path='task2_train_reformat.xlsx'):
    wb = load_workbook(path)
    ws = wb['sheet1']
    max_row = ws.max_row
    sum_line = 0
    first_sentence = 0
    second_sentence = 0
    third_sentence = 0
    four_sentence = 0
    five_sentence = 0
    for i in range(max_row - 1):
        row = i + 2
        if ws.cell(row, 2).value is not None:
            origins = ws.cell(row, 2).value.split(',')
            texts = ws.cell(row, 1).value.split('。')
            for origin in origins:
                sum_line += 1
                if origin in texts[0]:
                    first_sentence+=1
                if origin in texts[0] or origin in texts[1]:
                    second_sentence+=1
                if origin in texts[0] or origin in texts[1] or origin in texts[2]:
                    third_sentence+=1
                if origin in texts[0] or origin in texts[1] or origin in texts[2] or origin in texts[3]:
                    four_sentence+=1
                if origin in texts[0] or origin in texts[1] or origin in texts[2] or origin in texts[3] or origin in texts[4]:
                    five_sentence+=1
    print('origin出现在    第一句的概率：{}   前两句：{}  前三句：{}  前四句：{}  前五句：{}'.format(first_sentence/sum_line, second_sentence/sum_line, third_sentence/sum_line, four_sentence/sum_line, five_sentence/sum_line))

def analysis_conclusion(path='task2_train_reformat.xlsx'):
    wb = load_workbook(path)
    ws = wb['sheet1']
    max_row = ws.max_row
    num = 0
    for i in range(max_row - 1):
        row = i + 2
        text = ws.cell(row, 1).value
        flag = False
        if text.__contains__('1.') and text.__contains__('2.'):
            flag = True
        elif text.__contains__('1、') and text.__contains__('2、'):
            flag = True
        if flag:
            num+=1
    print('结论是固定格式的句子比例：{}'.format(num/(max_row-1)))

def trans_keyword_analysis(path='task2_train_reformat.xlsx'):
    wb = load_workbook(path)
    ws = wb['sheet1']
    max_row = ws.max_row
    contain = 0
    uncontain = 0
    for line in range(2, max_row+1):
        if ws.cell(line, 4).value is not None:
            if ws.cell(line, 1).value.__contains__('转移'):
                contain+=1
            else:
                print(ws.cell(line, 1).value, ws.cell(line, 4).value)
                uncontain+=1
    print('contian:{}   uncontain:{}'.format(contain, uncontain))

if __name__ == '__main__':
    # get_lengths('./sub_cut_train.xlsx')
    # tagging_notagging()
    # tags_proportion()
    # analysis_tag_in_sentence(path='task2_train_reformat_cleaned.xlsx')
    # origin_proportion_i_sentence()
    # analysis_conclusion(path='task2_train_reformat_cleaned.xlsx')
    trans_keyword_analysis()
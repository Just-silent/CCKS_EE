# -*- coding: utf-8 -*-
# @Author   : Just-silent
# @time     : 2020/6/26 15:56

import re
from tqdm import tqdm
from config import config
from random import shuffle
from tool import logger, tool
from openpyxl import load_workbook, Workbook

def sub_text(file='train'):
    if file == 'train':
        path = './sub_train.xlsx'
        save_path = './sub_cut_train.xlsx'
    else:
        path = './sub_dev.xlsx'
        save_path = './sub_cut_dev.xlsx'
    wb = load_workbook(path)
    ws = wb['sheet1']
    max_row = ws.max_row
    wb1 = Workbook()
    wb2 = Workbook()
    ws1 = wb1.create_sheet('sheet1')
    wb1.remove(wb1['Sheet'])
    ws2 = wb2.create_sheet('sheet1')
    wb2.remove(wb2['Sheet'])
    names = ['原文', '原发部位', '病灶大小', '转移部位']
    for i in range(len(names)):
        ws1.cell(1, i + 1, names[i])
        ws2.cell(1, i + 1, names[i])
    line_1_2 = 2
    for i in tqdm(range(max_row - 1)):
        p = [[] for i in range(6)]
        line = i + 2
        text = ws.cell(line,1).value
        middle = ws.cell(line,1).value.find('。')
        text1 = text[:middle+1]
        text2 = text[middle+1:]
        for i in range(3):
            if ws.cell(line,i+2).value is not None:
                places = ws.cell(line,i+2).value.split(',')
                for place in places:
                    if i==1:
                        p[2*i].append(place)
                        p[2 * i + 1].append(place)
                    else:
                        if place in text1:
                            p[2*i].append(place)
                        if place in text2:
                            p[2*i+1].append(place)
            else:
                p[2 * i]=''
                p[2 * i + 1]=''
        ws1.cell(line_1_2,1,text1)
        ws1.cell(line_1_2+1,1,text2)
        ws1.cell(line_1_2,2,','.join(p[0]))
        ws1.cell(line_1_2+1,2,','.join(p[1]))
        ws1.cell(line_1_2,3,','.join(p[2]))
        ws1.cell(line_1_2+1,3,','.join(p[3]))
        ws1.cell(line_1_2,4,','.join(p[4]))
        ws1.cell(line_1_2+1,4,','.join(p[5]))
        line_1_2+=2
    wb1.save(save_path)
    logger.info('Finished cut {}.xlsx'.format(file))

def sub_text_more(file='train'):
    if file == 'train':
        path = './sub_train.xlsx'
        save_path = './sub_cut_train.xlsx'
    else:
        path = './sub_dev.xlsx'
        save_path = './sub_cut_dev.xlsx'
    wb = load_workbook(path)
    ws = wb['sheet1']
    max_row = ws.max_row
    wb1 = Workbook()
    ws1 = wb1.create_sheet('sheet1')
    wb1.remove(wb1['Sheet'])
    names = ['原文', '原发部位', '病灶大小', '转移部位']
    for i in range(len(names)):
        ws1.cell(1, i + 1, names[i])
    all_text = []
    all_origin = []
    all_size = []
    all_trans = []
    for i in range(max_row-1):
        line = i+2
        text = ws.cell(line,1).value
        texts = tool.split_text(text)
        all_text.extend(texts)
        for j in range(3):
            if ws.cell(line,j+2).value is not None:
                places = ws.cell(line,j+2).value.split(',')
                for t in texts:
                    place_in_text = []
                    for place in places:
                        if place in t:
                            place_in_text.append(place)
                    if j==0:
                        all_origin.append(','.join(place_in_text))
                    elif j==1:
                        all_size.append(','.join(place_in_text))
                    else:
                        all_trans.append(','.join(place_in_text))
            else:
                for t in texts:
                    if j==0:
                        all_origin.append('')
                    elif j==1:
                        all_size.append('')
                    else:
                        all_trans.append('')
    assert len(all_trans) == len(all_size) and len(all_trans) == len(all_origin), 'len(all_trans) != len(all_size) or len(all_trans) != len(all_origin)'
    for i in range(len(all_text)):
        line = i+2
        ws1.cell(line,1,all_text[i])
        ws1.cell(line,2,all_origin[i])
        ws1.cell(line,3,all_size[i])
        ws1.cell(line,4,all_trans[i])
    wb1.save(save_path)
    logger.info('Finished cut {}.xlsx'.format(file))

def sub_text_condition(file='train'):
    if file == 'train':
        path = './sub_train.xlsx'
        save_path = './sub_cut_train1.xlsx'
    else:
        path = './sub_dev.xlsx'
        save_path = './sub_cut_dev1.xlsx'
    wb = load_workbook(path)
    ws = wb['sheet1']
    max_row = ws.max_row
    wb1 = Workbook()
    ws1 = wb1.create_sheet('sheet1')
    wb1.remove(wb1['Sheet'])
    names = ['原文', '原发部位', '病灶大小', '转移部位']
    for i in range(len(names)):
        ws1.cell(1, i + 1, names[i])
    all_text = []
    all_origin = []
    all_size = []
    all_trans = []
    for i in range(max_row-1):
        line = i+2
        text = ws.cell(line,1).value
        texts = tool.split_text(text)
        all_text.extend(texts)
        if ws.cell(line,4).value is not None and text.__contains__('转移'):
            places = ws.cell(line,4).value.split(',')
            for t in texts:
                place_in_text = []
                for place in places:
                    if place in t and t.__contains__('转移'):
                        place_in_text.append(place)
                all_trans.append(','.join(place_in_text))
        elif ws.cell(line,4).value is not None and not text.__contains__('转移'):
            places = ws.cell(line, 4).value.split(',')
            for t in texts:
                place_in_text = []
                for place in places:
                    if place in t :
                        place_in_text.append(place)
                all_trans.append(','.join(place_in_text))
        elif ws.cell(line,4).value is None:
            for t in texts:
                all_trans.append('')
        for j in range(2):
            if ws.cell(line,j+2).value is not None:
                places = ws.cell(line,j+2).value.split(',')
                for t in texts:
                    place_in_text = []
                    for place in places:
                        if place in t:
                            place_in_text.append(place)
                    if j==0:
                        all_origin.append(','.join(place_in_text))
                    elif j==1:
                        all_size.append(','.join(place_in_text))
                    else:
                        all_trans.append(','.join(place_in_text))
            else:
                for t in texts:
                    if j==0:
                        all_origin.append('')
                    elif j==1:
                        all_size.append('')
                    else:
                        all_trans.append('')
        # line = i+2
        # text = ws.cell(line, 1).value
        # describe, conclusion = tool.split_describe_conclusion(text)
        # origin = ws.cell(line, 2).value
        # size = ws.cell(line, 3).value
        # tran = ws.cell(line, 4).value
        # if conclusion is not None:
        #     conclusions = tool.split_text(conclusion)
        #     all_text.extend(conclusions)
        #     if origin is not None :
        #         origins = origin.split(',')
        #         for text in conclusions:
        #             place_in_text = []
        #             for place in origins:
        #                 if place in text:
        #                     place_in_text.append(place)
        #             all_origin.append(''.join(place_in_text))
        #     else:
        #         for i in range(len(conclusions)):
        #             all_origin.append('')
        #     if tran is not None:
        #         trans = tran.split(',')
        #         for text in conclusions:
        #             place_in_text = []
        #             for place in trans:
        #                 if place in text and place.__contains__('转移'):
        #                     place_in_text.append(place)
        #             all_trans.append(','.join(place_in_text))
        #     else:
        #         for i in range(len(conclusions)):
        #             all_trans.append('')
        #     for i in range(len(conclusions)):
        #         all_size.append('')
        #     describes = tool.split_text(describe)
        #     all_text.extend(describes)
        #     if origin is not None and size is not None:
        #         origins = origin.split(',')
        #         sizes = size.split(',')
        #         for text in describes:
        #             place_in_text1 = []
        #             place_in_text2 = []
        #             for place1 in origins:
        #                 for place2 in sizes:
        #                     if place1 in text and place2 in text:
        #                         place_in_text1.append(place1)
        #                         place_in_text2.append(place2)
        #             all_origin.append(','.join(place_in_text1))
        #             all_size.append(','.join(place_in_text2))
        #             all_trans.append('')
        #     elif origin is not None and size is None:
        #         origins = origin.split(',')
        #         for text in describes:
        #             place_in_text = []
        #             for place in origins:
        #                 if place in text:
        #                     place_in_text.append(place)
        #             all_origin.append(','.join(place_in_text))
        #             all_size.append('')
        #             all_trans.append('')
        #     else:
        #         for i in range(len(describes)):
        #             all_origin.append('')
        #             all_size.append('')
        #             all_trans.append('')
        # else:
        #     describes = tool.split_text(describe)
        #     all_text.extend(describes)
        #     if origin is not None and size is not None:
        #         origins = origin.split(',')
        #         sizes = size.split(',')
        #         for text in describes:
        #             place_in_text1 = []
        #             place_in_text2 = []
        #             for place1 in origins:
        #                 for place2 in sizes:
        #                     if place1 in text and place2 in text:
        #                         place_in_text1.append(place1)
        #                         place_in_text2.append(place2)
        #             all_origin.append(','.join(place_in_text1))
        #             all_size.append(','.join(place_in_text2))
        #     elif origin is not None and size is None:
        #         origins = origin.split(',')
        #         for text in describes:
        #             place_in_text = []
        #             for place in origins:
        #                 if place in text:
        #                     place_in_text.append(place)
        #             all_origin.append(','.join(place_in_text))
        #             all_size.append('')
        #     elif origin is None and size is None:
        #         for text in describes:
        #             all_origin.append('')
        #             all_size.append('')
        #     if tran is not None:
        #         trans = tran.split(',')
        #         for text in describes:
        #             place_in_text = []
        #             for place in trans:
        #                 if place in text and text.__contains__('转移'):
        #                     place_in_text.append(place)
        #             all_trans.append(','.join(place_in_text))
        #     else:
        #         for text in describes:
        #             all_trans.append('')

    assert len(all_text) == len(all_trans) and len(all_trans) == len(all_size) and len(all_trans) == len(all_origin), 'len(all_trans) != len(all_size) or len(all_trans) != len(all_origin)'
    for i in range(len(all_text)):
        line = i+2
        ws1.cell(line,1,all_text[i])
        ws1.cell(line,2,all_origin[i])
        ws1.cell(line,3,all_size[i])
        ws1.cell(line,4,all_trans[i])
    wb1.save(save_path)
    logger.info('Finished cut {}.xlsx'.format(file))

def data_clean(path='./task2_train_reformat{}.xlsx'):
    wb = load_workbook(path.format(''))
    ws = wb['sheet1']
    max_row = ws.max_row
    wb1 = Workbook()
    ws1 = wb1.create_sheet('sheet1')
    wb1.remove(wb1['Sheet'])
    names = ['原文', '原发部位', '病灶大小', '转移部位']
    for i in range(len(names)):
        ws1.cell(1, i + 1, names[i])
    place_num = 0
    size_num = 0
    for i in range(max_row - 1):
        line = i + 2
        new_sentence = ''
        chars = ['.','*','×','X','x','c','C','m','M',' ']
        # o_chars = ['_x0004_', '�', ':', ',', ';']
        # t_chars = ['', '', '：', '，', '；']
        o_chars = ['�']
        t_chars = ['']
        for i in range(4):
            if i==0:
                if '检测值' in ws.cell(line,i+1).value:
                    new_sentence = ws.cell(line, i + 1).value
                    for i in range(len(o_chars)):
                        new_sentence = new_sentence.replace(o_chars[i], t_chars[i])
                else:
                    new_sentence = ws.cell(line, i+1).value.replace(' ', '')
                    for i in range(len(o_chars)):
                        new_sentence = new_sentence.replace(o_chars[i], t_chars[i])
                i = 0
                j = 0
                while j < len(new_sentence):
                    while j < len(new_sentence) and not new_sentence[j].isdigit():
                        j+=1
                    start = j
                    end = start
                    while end+1<len(new_sentence) and (new_sentence[end+1] in chars or new_sentence[end+1].isdigit()):
                        end+=1
                    if new_sentence[start:end+1].__contains__('m') or new_sentence[start:end+1].__contains__('M'):
                        old_size = new_sentence[start:end+1]
                        new_size = ''
                        nums = ''
                        k=start
                        flag = False
                        while k<=end:
                            while new_sentence[k].isdigit() or new_sentence[k]=='.':
                                nums+=new_sentence[k]
                                k+=1
                                flag=True
                            if flag:
                                nums+=','
                                flag=False
                            k+=1
                        nums = nums[:-1].split(',')
                        if old_size.__contains__('c') or old_size.__contains__('C'):
                            for num in nums:
                                new_size = new_size+num+'CM'+'×'
                            new_size = new_size[:-1]
                        else:
                            for num in nums:
                                new_size = new_size+num+'MM'+'×'
                            new_size = new_size[:-1]
                        j=end
                        new_sentence = new_sentence.replace(old_size,new_size)
                    j+=1
                ws1.cell(line,i+1,new_sentence)
            elif i==1 or i==3:
                places = ws.cell(line,i+1).value
                if places is not None:
                    places = places.replace('_x0004_', '').replace(' ', '')
                    for place in places.split(','):
                        if place not in new_sentence:
                            place_num+=1
                            print('sentence中未找到place，次数{}  {}  {}'.format(place_num, new_sentence, place))
                    ws1.cell(line, i + 1, places)
                else:
                    ws1.cell(line, i + 1, '')
            else:
                sizes = ws.cell(line,i+1).value
                if sizes is not None:
                    sizes = sizes.replace('_x0004_', '').replace(' ', '').replace('�', '')
                    sizes_clean = ''
                    sizes_list = sizes.split(',')
                    for j in range(len(sizes_list)):
                        size = re.findall(r"\d+\.?\d*",sizes_list[j])
                        if sizes_list[j].__contains__('c') or sizes_list[j].__contains__('C'):
                            for k in range(len(size)):
                                sizes_clean = sizes_clean + size[k] + 'CM' + '×'
                            sizes_clean = sizes_clean[:-1]+','
                        else:
                            for k in range(len(size)):
                                sizes_clean = sizes_clean + size[k] + 'MM' + '×'
                            sizes_clean = sizes_clean[:-1] + ','
                    sizes_clean = sizes_clean[:-1]
                    for size in sizes_clean.split(','):
                        if size not in new_sentence:
                            size_num += 1
                            print('sentence中未找到place，次数{}  {}  {}'.format(size_num, new_sentence, size))
                    ws1.cell(line, i + 1, sizes_clean)
    wb1.save(path.format('_cleaned'))
    logger.info('Finished cleaned data')

def seg_train(path=config.train_dev_path):
        wb = load_workbook(path)
        ws = wb['sheet1']
        max_row = ws.max_row
        indexs = list(range(2,max_row+1))
        shuffle(indexs)
        wb_train = Workbook()
        sheet_train = wb_train.create_sheet('sheet1')
        wb_train.remove(wb_train['Sheet'])
        wb_dev = Workbook()
        sheet_dev = wb_dev.create_sheet('sheet1')
        wb_dev.remove(wb_dev['Sheet'])
        for i in range(4):
            sheet_train.cell(1,i+1,ws.cell(1,i+1).value)
            sheet_dev.cell(1,i+1,ws.cell(1,i+1).value)
        mid = len(indexs) // 10 * 8
        train_line = 2
        test_line = 2
        for i in range(len(indexs)):
            if i<mid:
                sheet_train.cell(train_line,1,ws.cell(indexs[i],1).value)
                sheet_train.cell(train_line,2,ws.cell(indexs[i],2).value)
                sheet_train.cell(train_line,3,ws.cell(indexs[i],3).value)
                sheet_train.cell(train_line,4,ws.cell(indexs[i],4).value)
                train_line+=1
            else:
                sheet_dev.cell(test_line,1,ws.cell(indexs[i],1).value)
                sheet_dev.cell(test_line,2,ws.cell(indexs[i],2).value)
                sheet_dev.cell(test_line,3,ws.cell(indexs[i],3).value)
                sheet_dev.cell(test_line,4,ws.cell(indexs[i],4).value)
                test_line+=1
        wb_train.save('./sub_train.xlsx')
        wb_dev.save('./sub_dev.xlsx')
        logger.info("Finished seg train data")

def data_clean_test(path='./task2_no_val{}.xlsx'):
    wb = load_workbook(path.format(''))
    ws = wb['sheet1']
    max_row = ws.max_row
    wb1 = Workbook()
    ws1 = wb1.create_sheet('sheet1')
    wb1.remove(wb1['Sheet'])
    names = ['原文', '肿瘤原发部位', '原发病灶大小', '转移部位']
    for i in range(len(names)):
        ws1.cell(1, i + 1, names[i])
    for i in range(max_row - 1):
        line = i + 2
        new_sentence = ''
        chars = ['.','*','×','X','x','c','C','m','M']
        # o_chars = ['_x0004_', '�', ':', ',', ';']
        # t_chars = ['', '', '：', '，', '；']
        o_chars = ['�']
        t_chars = ['']
        for i in range(4):
            if i==0:
                if '检测值' in ws.cell(line, i + 1).value:
                    new_sentence = ws.cell(line, i + 1).value
                    for j in range(len(o_chars)):
                        new_sentence = new_sentence.replace(o_chars[j], t_chars[j])
                else:
                    new_sentence = ws.cell(line, i+1).value.replace(' ', '')
                    # new_sentence = ws.cell(line, i + 1).value
                    for j in range(len(o_chars)):
                        new_sentence = new_sentence.replace(o_chars[j], t_chars[j])
                j = 0
                while j < len(new_sentence):
                    while j < len(new_sentence) and not new_sentence[j].isdigit():
                        j+=1
                    start = j
                    end = start
                    while end+1<len(new_sentence) and (new_sentence[end+1] in chars or new_sentence[end+1].isdigit()):
                        end+=1
                    if new_sentence[start:end+1].__contains__('m') or new_sentence[start:end+1].__contains__('M'):
                        old_size = new_sentence[start:end+1]
                        new_size = ''
                        nums = ''
                        k=start
                        flag = False
                        while k<=end:
                            while new_sentence[k].isdigit() or new_sentence[k]=='.':
                                nums+=new_sentence[k]
                                k+=1
                                flag=True
                            if flag:
                                nums+=','
                                flag=False
                            k+=1
                        nums = nums[:-1].split(',')
                        if old_size.__contains__('c') or old_size.__contains__('C'):
                            for num in nums:
                                new_size = new_size+num+'CM'+'×'
                            new_size = new_size[:-1]
                        else:
                            for num in nums:
                                new_size = new_size+num+'MM'+'×'
                            new_size = new_size[:-1]
                        j=end
                        new_sentence = new_sentence.replace(old_size,new_size)
                    j+=1
                ws1.cell(line,i+1,new_sentence)
    wb1.save(path.format('_cleaned'))
    logger.info('Finished val cleaned data')

def get_all_vocab():
    vocab_train_path = './task2_vocab.txt'
    vocab_val_path = './task2_vocab.val.txt'
    train_data_path = './task2_train_reformat.xlsx'
    all_cocab_path = './all_vocab.txt'
    vocab_list = []
    for path in [vocab_train_path, vocab_val_path]:
        with open(path, 'r', encoding='utf-8') as vocab:
            for x in vocab.readlines():
                vocab_list.append(x[:-1].replace('_x0004_', '').replace(' ', ''))
    wb = load_workbook(train_data_path)
    ws = wb['sheet1']
    max_row = ws.max_row
    for i in range(max_row-1):
        line = i+2
        if ws.cell(line, 2).value is not None:
            places = ws.cell(line, 2).value.split(',')
            for place in places:
                vocab_list.append(place.replace('_x0004_', '').replace(' ', ''))
        if ws.cell(line, 4).value is not None:
            places = ws.cell(line, 4).value.split(',')
            for place in places:
                vocab_list.append(place.replace('_x0004_', '').replace(' ', ''))
    vocab_list = list(set(vocab_list))
    with open(all_cocab_path, 'w', encoding='utf-8') as all_cocab_file:
        for place in vocab_list:
            all_cocab_file.write(place+'\n')
    logger.info('all_vocab写入完成')

def get_three_files(file='train'):
    if file == 'train':
        path = './sub_train.xlsx'
        save_conclusion_path = './conclusion_train.xlsx'
        save_describe_path = './describe_train.xlsx'
        save_others_path = './others_train.xlsx'
    else:
        path = './sub_dev.xlsx'
        save_conclusion_path = './conclusion_dev.xlsx'
        save_describe_path = './describe_dev.xlsx'
        save_others_path = './others_dev.xlsx'
    wb = load_workbook(path)
    ws = wb['sheet1']
    max_row = ws.max_row
    wb1 = Workbook()
    ws1 = wb1.create_sheet('sheet1')
    wb1.remove(wb1['Sheet'])
    wb2 = Workbook()
    ws2 = wb2.create_sheet('sheet1')
    wb2.remove(wb2['Sheet'])
    wb3 = Workbook()
    ws3 = wb3.create_sheet('sheet1')
    wb3.remove(wb3['Sheet'])
    names = ['原文', '原发部位', '病灶大小', '转移部位']
    line1 = 2
    line2 = 2
    line3 = 2
    for i in range(len(names)):
        ws1.cell(1, i + 1, names[i])
        ws2.cell(1, i + 1, names[i])
        ws3.cell(1, i + 1, names[i])
    for row in tqdm(range(2,max_row + 1)):
        new_origin_places = ''
        new_size = ''
        new_trans = ''
        text = ws.cell(row, 1).value
        describe, conclusion = tool.split_describe_conclusion(text)
        if conclusion is not None:
            origin_places = ws.cell(row, 2).value
            sizes= ws.cell(row, 3).value
            transfered_places = ws.cell(row, 4).value
            for i, place_str in enumerate([origin_places, sizes, transfered_places]):
                if place_str is not None:
                    if i==2:
                        new_place = ''
                        places = place_str.split(',')
                        for place in places:
                            if place in conclusion:
                                new_place = new_place + place + ','
                        new_trans = new_place[:-1]
                    elif i==1:
                        new_place = ''
                        places = place_str.split(',')
                        for place in places:
                            if place in describe:
                                new_place = new_place + place + ','
                        new_size = new_place[:-1]
                    else:
                        new_place = ''
                        places = place_str.split(',')
                        for place in places:
                            if place in describe:
                                new_place = new_place + place + ','
                        new_origin_places = new_place[:-1]
                else:
                    if i==2:
                        new_trans = None
                    elif i==1:
                        new_size = None
                    else:
                        new_origin_places = None
            ws1.cell(line1, 1, describe)
            ws1.cell(line1, 2, new_origin_places)
            ws1.cell(line1, 3, new_size)
            ws2.cell(line2, 1, conclusion)
            ws2.cell(line2, 4, new_trans)
            line1+=1
            line2+=1
        else:
            origin_places = ws.cell(row, 2).value
            sizes = ws.cell(row, 3).value
            transfered_places = ws.cell(row, 4).value
            ws3.cell(line3, 1, describe)
            ws3.cell(line3, 2, origin_places)
            ws3.cell(line3, 3, sizes)
            ws3.cell(line3, 4, transfered_places)
            line3+=1
    wb1.save(save_describe_path)
    wb2.save(save_conclusion_path)
    wb3.save(save_others_path)

if __name__ == '__main__':
    # data_clean()
    # seg_train('./task2_train_reformat_cleaned.xlsx')
    files = ['train','dev']
    for file in files:
        # sub_text_more(file)
        # get_three_files(file)
        sub_text_condition(file)
    # data_clean_test()
    # get_all_vocab()

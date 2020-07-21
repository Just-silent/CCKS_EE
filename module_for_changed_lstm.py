# -*- coding: utf-8 -*-
# @Author   : Just-silent
# @time     : 2020/7/13 21:36

# -*- coding: utf-8 -*-
# @Author   : Just-silent
# @time     : 2020/4/26 10:24

import os
import re
import numpy
import torch
import random
from tqdm import tqdm
from config import config, device
from tool import tool, logger, get_all_tag
from openpyxl import load_workbook
from model import TransformerEncoderModel, BiLSTM_CRF, CNN_CRF, BiLSTM_CRF_ATT, BiLSTM_CRF_changed, TransformerEncoderModel_DAE
from sklearn.metrics import classification_report
from result.predict_eval_process import format_result
import torch.optim as optim

seed = 0
torch.manual_seed(seed)
torch.cuda.manual_seed(seed)
torch.cuda.manual_seed_all(seed)  # if you are using multi-GPU.
numpy.random.seed(seed)  # Numpy module.
random.seed(seed)  # Python random module.
torch.manual_seed(seed)
torch.backends.cudnn.benchmark = False
torch.backends.cudnn.deterministic = True


import warnings
warnings.filterwarnings('ignore')

def data_loader(path, sort=True):
    data = []
    wb = load_workbook(filename=path)
    ws = wb['sheet1']
    max_row = ws.max_row
    for line_num in range(max_row - 1):
        line_num = line_num + 2
        sentence, origin_places, sizes, transfered_places = ws.cell(line_num, 1).value, ws.cell(line_num,
                                                                                                2).value, ws.cell(
            line_num, 3).value, ws.cell(line_num, 4).value
        if sentence is not None:
            tag_list = get_all_tag(sentence, origin_places, sizes, transfered_places)
            texts = re.split('。',sentence)
            sub_len = 0
            if '' in texts:
                sub_len = len(texts)-1
            else:
                sub_len = len(texts)
            sentence_list = [x for x in sentence]
            data.append({'text':sentence_list, 'tag':tag_list, 'sub_len':sub_len})
    data = sorted(data, key=lambda x:x['sub_len'], reverse=True)
    return data

def build_stoi_itos(train_data, dev_data, kind):
    all_char = {}
    stoi = {}
    itos = {}
    for data in train_data:
        text = data[kind]
        for c in text:
            if c not in all_char.keys():
                all_char[c] = 1
            else:
                all_char[c] += 1
    for data in dev_data:
        text = data[kind]
        for c in text:
            if c not in all_char.keys():
                all_char[c] = 1
            else:
                all_char[c] += 1
    all_char = dict(sorted(all_char.items(), key=lambda x: x[1], reverse=True))
    stoi = {'unk': 0, 'pad': 1}
    i = 2
    for c in all_char.keys():
        stoi[c] = i
        i += 1
    itos = dict(zip(stoi.values(), stoi.keys()))
    return stoi, itos

def get_batch(data, text_stoi, tag_stoi, batch_size):
    iter = {'text':[], 'tag':[], 'sub_tag':[], 'text_len':[]}
    chunk = len(data)//batch_size if len(data)%batch_size==0 else len(data)//batch_size+1
    for num in range(chunk):
        iter_i = {'text':[], 'tag':[], 'sub_tag':[], 'text_len':[]}
        max_sub_len = 0
        if num!=chunk-1:
            datas = data[num*batch_size:(num+1)*batch_size]
        else:
            datas = data[num * batch_size:]
        max_sub = datas[0]['sub_len']
        for data_i in datas:
            texts_o = []
            sub_tag = []
            text_len = []
            text = ''.join(data_i['text'])
            tag = data_i['tag']
            texts = re.split('。', text)
            for text in texts:
                if '' in texts:
                    if text!='':
                        texts_o.append(text+'。')
                else:
                    if text!=texts[-1:]:
                        texts_o.append(text+'。')
                    else:
                        texts_o.append(text)
            for i in range(len(texts_o)):
                texts_o[i] = [c for c in texts_o[i]]
            start = 0
            for text in texts_o:
                sub_tag.append(tag[start:start+len(text)])
                start+=len(text)
                text_len.append(len(text))
                if len(text)>max_sub_len:
                    max_sub_len = len(text)
            iter_i['text'].append(texts_o)
            iter_i['tag'].append(tag)
            iter_i['sub_tag'].append(sub_tag)
            iter_i['text_len'].append(text_len)
        for j in range(batch_size):
            if len(iter_i['text'])<batch_size:
                for i in range(batch_size-len(iter_i['text'])):
                    texts_o = [['pad' for k in range(max_sub_len)] for num in range(max_sub)]
                    sub_tag_o = [['pad' for k in range(max_sub_len)] for num in range(max_sub)]
                    tags_o = ['pad' for k in range(max_sub_len*max_sub)]
                    text_len = [0 for num in range(max_sub)]
                    iter_i['text'].append(texts_o)
                    iter_i['tag'].append(tags_o)
                    iter_i['sub_tag'].append(sub_tag_o)
                    iter_i['text_len'].append(text_len)
            if len(iter_i['tag'][j])<max_sub_len*max_sub:
                iter_i['tag'][j].extend(['pad' for i in range(max_sub_len*max_sub-len(iter_i['tag'][j]))])
            for i in range(len(iter_i['text'][j])):
                if len(iter_i['text'][j][i])<max_sub_len:
                    iter_i['text'][j][i].extend(['pad' for i in range(max_sub_len-len(iter_i['text'][j][i]))])
                if len(iter_i['sub_tag'][j][i]) < max_sub_len:
                    iter_i['sub_tag'][j][i].extend(['pad' for i in range(max_sub_len-len(iter_i['sub_tag'][j][i]))])
            if len(iter_i['text'][j])<max_sub:
                for k in range(max_sub-len(iter_i['text'][j])):
                    iter_i['text'][j].append(['pad' for i in range(max_sub_len)])
                    iter_i['sub_tag'][j].append(['pad' for i in range(max_sub_len)])
                    iter_i['text_len'][j].append(0)
        for i in range(batch_size):
            iter_i['tag'][i] = [tag_stoi[c] for c in iter_i['tag'][i]]
            for j in range(len(iter_i['text'][i])):
                iter_i['text'][i][j] = [text_stoi[c] for c in iter_i['text'][i][j]]
                iter_i['sub_tag'][i][j] = [tag_stoi[c] for c in iter_i['sub_tag'][i][j]]
        text_tensor = torch.tensor(numpy.array(iter_i['text'], dtype=numpy.int64)).to(device)
        tag_tensor = torch.tensor(numpy.array(iter_i['tag'], dtype=numpy.int64)).to(device)
        sub_tag_tensor = torch.tensor(numpy.array(iter_i['sub_tag'], dtype=numpy.int64)).to(device)
        text_len_tensor = torch.tensor(numpy.array(iter_i['text_len'], dtype=numpy.int64)).to(device)
        iter['text'].append(text_tensor)
        iter['tag'].append(tag_tensor)
        iter['sub_tag'].append(sub_tag_tensor)
        iter['text_len'].append(text_len_tensor)
    return iter

def get_batch_sub_text(data, text_stoi, tag_stoi, batch_size):
    iter = {'text':[], 'tag':[], 'sub_tag':[], 'hidden_tag':[], 'text_len':[]}
    chunk = len(data)//batch_size if len(data)%batch_size==0 else len(data)//batch_size+1
    for num in range(chunk):
        iter_i = {'text':[], 'tag':[], 'sub_tag':[], 'hidden_tag':[], 'text_len':[]}
        max_sub_len = 0
        if num!=chunk-1:
            datas = data[num*batch_size:(num+1)*batch_size]
        else:
            datas = data[num * batch_size:]
        max_sub = datas[0]['sub_len']
        for data_i in datas:
            texts_o = []
            sub_tag = []
            text_len = []
            hidden_tag = []
            text = ''.join(data_i['text'])
            tag = data_i['tag']
            texts = re.split('。', text)
            for text in texts:
                if '' in texts:
                    if text!='':
                        texts_o.append(text+'。')
                else:
                    if text!=texts[-1:]:
                        texts_o.append(text+'。')
                    else:
                        texts_o.append(text)
            for i in range(len(texts_o)):
                texts_o[i] = [c for c in texts_o[i]]
            start = 0
            for text in texts_o:
                sub_tag.append(tag[start:start+len(text)])
                hidden_tag.append(is_hidden_tag(tag[start:start+len(text)]))
                start+=len(text)
                text_len.append(len(text))
                if len(text)>max_sub_len:
                    max_sub_len = len(text)
            iter_i['text'].append(texts_o)
            iter_i['tag'].append(tag)
            iter_i['sub_tag'].append(sub_tag)
            iter_i['hidden_tag'].append(hidden_tag)
            iter_i['text_len'].append(text_len)
        for j in range(batch_size):
            if len(iter_i['text'])<batch_size:
                for i in range(batch_size-len(iter_i['text'])):
                    texts_o = [['pad' for k in range(max_sub_len)] for num in range(max_sub)]
                    sub_tag_o = [['pad' for k in range(max_sub_len)] for num in range(max_sub)]
                    tags_o = ['pad' for k in range(max_sub_len*max_sub)]
                    text_len = [0 for num in range(max_sub)]
                    hidden_tag_o = [0 for num in range(max_sub)]
                    iter_i['text'].append(texts_o)
                    iter_i['tag'].append(tags_o)
                    iter_i['sub_tag'].append(sub_tag_o)
                    iter_i['hidden_tag'].append(hidden_tag_o)
                    iter_i['text_len'].append(text_len)
            if len(iter_i['tag'][j])<max_sub_len*max_sub:
                iter_i['tag'][j].extend(['pad' for i in range(max_sub_len*max_sub-len(iter_i['tag'][j]))])
            for i in range(len(iter_i['text'][j])):
                if len(iter_i['text'][j][i])<max_sub_len:
                    iter_i['text'][j][i].extend(['pad' for i in range(max_sub_len-len(iter_i['text'][j][i]))])
                if len(iter_i['sub_tag'][j][i]) < max_sub_len:
                    iter_i['sub_tag'][j][i].extend(['pad' for i in range(max_sub_len-len(iter_i['sub_tag'][j][i]))])
            if len(iter_i['text'][j])<max_sub:
                for k in range(max_sub-len(iter_i['text'][j])):
                    iter_i['text'][j].append(['pad' for i in range(max_sub_len)])
                    iter_i['sub_tag'][j].append(['pad' for i in range(max_sub_len)])
                    iter_i['hidden_tag'][j].append(0)
                    iter_i['text_len'][j].append(0)
        for i in range(batch_size):
            iter_i['tag'][i] = [tag_stoi[c] for c in iter_i['tag'][i]]
            for j in range(len(iter_i['text'][i])):
                iter_i['text'][i][j] = [text_stoi[c] for c in iter_i['text'][i][j]]
                iter_i['sub_tag'][i][j] = [tag_stoi[c] for c in iter_i['sub_tag'][i][j]]
        text_tensor = torch.tensor(numpy.array(iter_i['text'], dtype=numpy.int64)).to(device)
        tag_tensor = torch.tensor(numpy.array(iter_i['tag'], dtype=numpy.int64)).to(device)
        sub_tag_tensor = torch.tensor(numpy.array(iter_i['sub_tag'], dtype=numpy.int64)).to(device)
        hidden_tag_tensor = torch.tensor(numpy.array(iter_i['hidden_tag'], dtype=numpy.int64)).to(device)
        text_len_tensor = torch.tensor(numpy.array(iter_i['text_len'], dtype=numpy.int64)).to(device)
        iter['text'].append(text_tensor)
        iter['tag'].append(tag_tensor)
        iter['sub_tag'].append(sub_tag_tensor)
        iter['hidden_tag'].append(hidden_tag_tensor)
        iter['text_len'].append(text_len_tensor)
    return iter

def is_hidden_tag(tags):
    hidden_tag = 0
    for tag in tags:
        if tag != 'O':
            hidden_tag = 1
            break
    return hidden_tag

class EE():
    def __init__(self):
        self.model = None
        self.text_itos = None
        self.tag_itos = None
        self.train_dev_data = None

    def train(self):
        max_f1 = -1
        max_dict = {}
        max_report = {}
        loss_list = []
        f1_list = []
        epoch_list = []
        if not os.path.exists('./result/classification_report/{}'.format(config.experiment_name)):
            os.mkdir('./result/classification_report/{}'.format(config.experiment_name))
            os.mkdir('./result/picture/{}'.format(config.experiment_name))
            os.mkdir('./result/data/{}'.format(config.experiment_name))
            os.mkdir('./result/data/{}/test_format'.format(config.experiment_name))
        logger.info('Loading data ...')
        train_data= data_loader(config.train_path, sort=True)
        dev_data = data_loader(config.dev_path, sort=True)
        logger.info('Finished load data')
        logger.info('Building vocab ...')
        text_stoi, text_itos = build_stoi_itos(train_data, dev_data, 'text')
        tag_stoi, tag_itos = build_stoi_itos(train_data, dev_data, 'tag')
        self.tag_itos = tag_itos
        self.text_itos = text_itos
        logger.info('Finished build vocab')
        logger.info('Building iterator ...')
        train_iter = get_batch_sub_text(train_data, text_stoi, tag_stoi, batch_size=config.batch_size)
        dev_iter = get_batch_sub_text(dev_data, text_stoi, tag_stoi, batch_size=config.batch_size)
        logger.info('Finished build iterator')
        if config.model_name == 'BiLSTM_CRF_changed':
            model = BiLSTM_CRF_changed(config, ntoken=len(text_stoi), ntag=len(tag_stoi)).to(device)
        self.model = model
        optimizer = optim.Adam(model.parameters(), lr=config.learning_rate, weight_decay=1e-5)
        logger.info('Begining train ...')
        for epoch in range(config.epoch):
            model.train()
            acc_loss = 0
            for i in tqdm(range(len(train_iter['text']))):
                optimizer.zero_grad()
                text = train_iter['text'][i]
                tag = train_iter['tag'][i]
                sub_tag = train_iter['sub_tag'][i]
                hidden_tag = train_iter['hidden_tag'][i]
                text_len = train_iter['text_len'][i]
                loss = model.loss(text, text_len, tag, sub_tag, hidden_tag)
                acc_loss += loss.view(-1).cpu().data.tolist()[0]
                loss.backward()
                optimizer.step()
            f1, report_dict, entity_prf_dict = self.eval(dev_iter)
            loss_list.append(acc_loss)
            f1_list.append(f1)
            epoch_list.append(epoch+1)
            logger.info('epoch:{}   loss:{}   weighted avg:{}'.format(epoch, acc_loss, report_dict['weighted avg']))
            if f1 > max_f1:
                max_f1 = f1
                max_dict = entity_prf_dict['average']
                max_report = entity_prf_dict
                torch.save(model.state_dict(), './save_model/{}.pkl'.format(config.experiment_name))
        logger.info('Finished train')
        logger.info('Max_f1 avg : {}'.format(max_dict))
        tool.write_csv(max_report)
        tool.show_1y(epoch_list, loss_list, 'loss')
        tool.show_1y(epoch_list, f1_list, 'f1')
        # 稍后处理
        # tool.show_labels_f1_bar_divide(max_report)

    def eval(self, dev_iter):
        self.model.eval()
        tag_true_all = []
        tag_pred_all = []
        entity_prf_dict = {}
        model = self.model
        entities_total = {'origin_place': {'TP': 0, 'S': 0, 'G': 0, 'p': 0, 'r': 0, 'f1': 0},
                          'size': {'TP': 0, 'S': 0, 'G': 0, 'p': 0, 'r': 0, 'f1': 0},
                          'transfered_place': {'TP': 0, 'S': 0, 'G': 0, 'p': 0, 'r': 0, 'f1': 0}}
        for i in tqdm(range(len(dev_iter['text']))):
            text = dev_iter['text'][i]
            tag = dev_iter['tag'][i]
            text_len = dev_iter['text_len'][i]
            text_len1 = [sum(dev_iter['text_len'][i][j]) for j in range(dev_iter['text_len'][i].size(0))]
            tag_list = []
            for i in range(config.batch_size):
                tag_list.extend(tag[i][:text_len1[i]].cpu().numpy())
            result = model(text, text_len)
            result_list = []
            for l in result:
                result_list.extend(l)
            assert len(tag_list) == len(result_list), 'tag_list: {} != result_list: {}'.format(len(tag_list),
                                                                                               len(result_list))
            tag_true = [self.tag_itos[k] for k in tag_list]
            tag_true_all.extend(tag_true)
            tag_pred = [self.tag_itos[k] for k in result_list]
            tag_pred_all.extend(tag_pred)
            entities = self._evaluate(tag_true=tag_true, tag_pred=tag_pred)
            assert len(entities_total) == len(entities), 'entities_total: {} != entities: {}'.format(
                len(entities_total), len(entities))
            for entity in entities_total:
                entities_total[entity]['TP'] += entities[entity]['TP']
                entities_total[entity]['S'] += entities[entity]['S']
                entities_total[entity]['G'] += entities[entity]['G']
        TP = 0
        S = 0
        G = 0
        print('\n--------------------------------------------------')
        print('\tp\t\t\tr\t\t\tf1\t\t\tlabel_type')
        for entity in entities_total:
            entities_total[entity]['p'] = entities_total[entity]['TP'] / entities_total[entity]['S'] \
                if entities_total[entity]['S'] != 0 else 0
            entities_total[entity]['r'] = entities_total[entity]['TP'] / entities_total[entity]['G'] \
                if entities_total[entity]['G'] != 0 else 0
            entities_total[entity]['f1'] = 2 * entities_total[entity]['p'] * entities_total[entity]['r'] / \
                                           (entities_total[entity]['p'] + entities_total[entity]['r']) \
                if entities_total[entity]['p'] + entities_total[entity]['r'] != 0 else 0
            print('\t{:.3f}\t\t{:.3f}\t\t{:.3f}\t\t{}'.format(entities_total[entity]['p'], entities_total[entity]['r'],
                                                              entities_total[entity]['f1'], entity))
            entity_dict = {'precision':entities_total[entity]['p'], 'recall':entities_total[entity]['r'], 'f1-score':entities_total[entity]['f1'], 'support':''}
            entity_prf_dict[entity] = entity_dict
            TP += entities_total[entity]['TP']
            S += entities_total[entity]['S']
            G += entities_total[entity]['G']
        p = TP / S if S != 0 else 0
        r = TP / G if G != 0 else 0
        f1 = 2 * p * r / (p + r) if p + r != 0 else 0
        print('\t{:.3f}\t\t{:.3f}\t\t{:.3f}\t\taverage'.format(p, r, f1))
        print('--------------------------------------------------')
        entity_prf_dict['average'] = {'precision':p, 'recall':r, 'f1-score':f1, 'support':''}
        labels = []
        for index, label in enumerate(self.tag_itos.values()):
            labels.append(label)
        labels.remove('O')
        prf_dict = classification_report(tag_true_all, tag_pred_all, labels=labels, output_dict=True)
        return f1, prf_dict, entity_prf_dict

    def _evaluate(self, tag_true, tag_pred):
        """
        先对true进行还原成 [{}] 再对pred进行还原成 [{}]
        :param tag_true: list[]
        :param tag_pred: list[]
        :return:
        """
        true_list = self._build_list_dict(_len=len(tag_true), _list=tag_true)
        pred_list = self._build_list_dict(_len=len(tag_pred), _list=tag_pred)
        entities = {'origin_place': {'TP': 0, 'S': 0, 'G': 0},
                    'size': {'TP': 0, 'S': 0, 'G': 0},
                    'transfered_place': {'TP': 0, 'S': 0, 'G': 0}}
        for true in true_list:
            label_type = true['label_type']
            entities[label_type]['G'] += 1
        for pred in pred_list:
            start_pos = pred['start_pos']
            end_pos = pred['end_pos']
            label_type = pred['label_type']
            entities[label_type]['S'] += 1
            for true in true_list:
                if label_type == true['label_type'] and start_pos == true['start_pos'] and end_pos == true['end_pos']:
                    entities[label_type]['TP'] += 1
        return entities

    def _build_list_dict(self, _len, _list):
        build_list = []
        tag_dict = {'origin_place': 'origin_place',
                    'size': 'size',
                    'transfered_place': 'transfered_place'}
        for index, tag in zip(range(_len), _list):
            if tag[0] == 'B':
                label_type = tag[2:]
                start_pos = index
                if index < _len-1:
                    end_pos = index + 1
                    while _list[end_pos][0] == 'I' and _list[end_pos][2:] == label_type and end_pos<_len-1:
                        end_pos += 1
                else:
                    end_pos = index
                build_list.append({'start_pos': start_pos,
                                   'end_pos': end_pos,
                                   'label_type': tag_dict[label_type]})
        return build_list

    def predict_test(self, path=config.test_path, model_name=config.model_path.format(config.experiment_name),  save_path=config.unformated_val_path.format(config.experiment_name)):
        train_data = tool.load_data(config.train_path)
        dev_data = tool.load_data(config.dev_path)
        logger.info('Finished load data')
        logger.info('Building vocab ...')
        model = None
        if config.is_pretrained_model:
            with open(config.pretrained_vocab, 'r', encoding='utf-8') as vocab_file:
                vocab_list = vocab_file.readlines()
            word_vocab = tool.get_text_vocab(vocab_list)
        else:
            word_vocab = tool.get_text_vocab(train_data, dev_data)
        vectors = word_vocab.vectors
        tag_vocab = tool.get_tag_vocab(train_data, dev_data)
        logger.info('Finished build vocab')
        if config.model_name == 'BiLSTM_CRF_changed':
            model = BiLSTM_CRF_changed(config, ntoken=len(self.word_vocab), ntag=len(self.tag_vocab),
                                       vectors=vectors).to(device)
        elif config.model_name == 'TransformerEncoderModel':
            model = TransformerEncoderModel(config, ntoken=len(self.word_vocab), ntag=len(self.tag_vocab),
                                            vectors=vectors).to(device)
        elif config.model_name == 'TransformerEncoderModel_DAE':
            model = TransformerEncoderModel_DAE(config, ntoken=len(self.word_vocab), ntag=len(self.tag_vocab),
                                                vectors=vectors).to(device)
        elif config.model_name == 'CNN_CRF':
            model = CNN_CRF(config, ntoken=len(self.word_vocab), ntag=len(self.tag_vocab)).to(device)
        elif config.model_name == 'BiLSTM_CRF_ATT':
            model = BiLSTM_CRF_ATT(config, ntoken=len(self.word_vocab), ntag=len(self.tag_vocab), vectors=vectors).to(
                device)
        model.load_state_dict(torch.load(model_name))
        wb = load_workbook(filename=path)
        ws = wb['sheet1']
        max_row = ws.max_row
        for line_num in tqdm(range(max_row-1)):
            line_num+=2
            sentence = ws.cell(line_num,1).value
            texts = re.split('。', sentence)
            tensors = []
            text_o = []
            text_len = 0
            for index, text in enumerate(texts):
                if '' in texts:
                    if text != '':
                        text_o.append(text + '。')
                else:
                    if index != len(texts) - 1:
                        text_o.append(text + '。')
                    else:
                        text_o.append(text)
            for text in text_o:
                tensors.append(
                    torch.tensor(numpy.array([self.word_vocab.stoi[s] for s in text], dtype='int64')).to(device))
            tag_pred = [list1[0] for list1 in model(tensors, text_len)]
            # middle = sentence.find('。')
            # sentence1 = sentence[:middle+1]
            # sentence2 = sentence[middle+1:]
            # text1 = torch.tensor(numpy.array([word_vocab.stoi[word] for word in sentence1], dtype='int64')).unsqueeze(1).expand(len(sentence1),config.batch_size).to(device)
            # text1_len = torch.tensor(numpy.array([len(sentence1)], dtype='int64')).expand(config.batch_size).to(device)
            # text2 = torch.tensor(numpy.array([word_vocab.stoi[word] for word in sentence2], dtype='int64')).unsqueeze(
            #     1).expand(len(sentence2), config.batch_size).to(device)
            # text2_len = torch.tensor(numpy.array([len(sentence2)], dtype='int64')).expand(config.batch_size).to(device)
            # result1 = model(text1, text1_len)[0]
            # result2 = model(text2, text2_len)[0]
            # tag_pred = []
            # for result in [result1, result2]:
            #     for k in result:
            #         tag_pred.append(tag_vocab.itos[k])

            i = 0
            origin_places = []
            sizes = []
            transfered_places = []
            # 此处可能存在问题
            while i < len(tag_pred):
                start = 0
                end = 0
                kind = None
                if tag_pred[i]!='O':
                    start = i
                    kind = tag_pred[i][2:]
                    while i+1<len(tag_pred) and tag_pred[i+1][2:]==kind:
                        i+=1
                    end = i + 1
                    if kind == 'origin_place':
                        origin_places.append(sentence[start:end])
                    elif kind == 'size':
                        sizes.append(sentence[start:end])
                    else:
                        transfered_places.append(sentence[start:end])
                i+=1
            ws.cell(line_num, 2).value = ','.join(list(set(origin_places)))
            ws.cell(line_num, 3).value = ','.join(list(set(sizes)))
            ws.cell(line_num, 4).value = ','.join(list(set(transfered_places)))
        wb.save(filename=save_path)
        logger.info('Finished Predicting...')

    def test_format_result(self):
        self.train_data = tool.load_data(config.train_path)
        self.dev_data = tool.load_data(config.dev_path)
        tag_vocab = tool.get_tag_vocab(self.train_data, self.dev_data)
        self.predict_test(path=config.dev_path, save_path=config.test_unformated_val_path.format(config.experiment_name))
        tag_true = []
        tag_formated_pred = []
        tag_unformated_pred = []
        format_result(path=config.test_unformated_val_path.format(config.experiment_name), save_path=config.test_formated_val_path.format(config.experiment_name))
        dev_data = tool.load_data(config.dev_path)
        formated_dev_data = tool.load_data(config.test_formated_val_path.format(config.experiment_name))
        unformated_dev_data = tool.load_data(config.test_unformated_val_path.format(config.experiment_name))
        assert len(dev_data.examples) == len(
            unformated_dev_data.examples), 'train_dev_data:{} != unformated_train_dev_data:{}'.format(
            len(dev_data.examples), len(unformated_dev_data.examples))
        assert len(dev_data.examples) == len(
            formated_dev_data.examples), 'train_dev_data:{} != formated_train_dev_data:{}'.format(
            len(dev_data.examples), len(formated_dev_data.examples))
        for example1 in dev_data.examples:
            tag_true.extend(example1.tag)
        for example2 in formated_dev_data.examples:
            tag_formated_pred.extend(example2.tag)
        for example3 in unformated_dev_data.examples:
            tag_unformated_pred.extend(example3.tag)
        # the eval of unformated result
        for i in range(len(dev_data)):
            pass
        assert len(tag_true) == len(tag_unformated_pred), 'tag_true:{} != tag_pred:{}'.format(len(tag_true), len(tag_unformated_pred))
        assert len(tag_true) == len(tag_formated_pred), 'tag_true:{} != tag_pred:{}'.format(len(tag_true), len(tag_formated_pred))
        labels = []
        for index, label in enumerate(tag_vocab.itos):
            labels.append(label)
        labels.remove('O')
        prf_dict_formated = classification_report(tag_true, tag_formated_pred, labels=labels, output_dict=True)
        prf_dict_unformated = classification_report(tag_true, tag_unformated_pred, labels=labels, output_dict=True)
        # the eval of formated result
        logger.info('unformated report{}'.format(prf_dict_formated['weighted avg']))
        logger.info('formated report{}'.format(prf_dict_unformated['weighted avg']))

if __name__ == '__main__':
    ee = EE()
    ee.train()
    # 可单独运行
    ee.predict_test()
    # 可单独运行
    # ee.test_format_result()
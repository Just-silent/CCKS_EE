# -*- coding: utf-8 -*-
# @Author   : Just-silent
# @time     : 2020/4/26 10:24
import codecs
import os
import re
import torch
import numpy
import random
from tqdm import tqdm
from config import device, config
from tool import Tool, logger, get_bigram, Trie
from openpyxl import load_workbook
from model import TransformerEncoderModel, BiLSTM_CRF, BiLSTM_CRF_hidden_tag, CNN_CRF, BiLSTM_CRF_ATT, \
    CNN_TransformerEncoderModel, TransformerEncoderModel_DAE, BiLSTM_CRF_DAE, FLAT
from sklearn.metrics import classification_report
from result.predict_eval_process import format_result
import torch.optim as optim

seed = 0
torch.manual_seed(seed)
torch.cuda.manual_seed(seed)
torch.cuda.manual_seed_all(seed)  # if you are using multi-GPU.
numpy.random.seed(seed)  # Numpy module.
random.seed(seed)  # Python random module.
torch.manual_seed(seed)
torch.backends.cudnn.benchmark = False
torch.backends.cudnn.deterministic = True

import warnings

warnings.filterwarnings('ignore')


class EE():
    def __init__(self, config):
        self.config = config
        self.model = None
        self.word_vocab = None
        self.tag_vocab = None
        self.train_dev_data = None
        self.bigram_vocab = None
        self.lattice_vocab = None
        self.tool = Tool(self.config)

    def init_model(self, config=None, ntoken=None, ntag=None, hidden_ntag=None, vectors=None, n_bigram=None):
        model_name = config.model_name
        models = {
            'CNN_CRF': CNN_CRF,
            'BiLSTM_CRF': BiLSTM_CRF,
            'BiLSTM_CRF_ATT': BiLSTM_CRF_ATT,
            'BiLSTM_CRF_DAE': BiLSTM_CRF_DAE,
            'BiLSTM_CRF_hidden_tag': BiLSTM_CRF_hidden_tag,
            'TransformerEncoderModel': TransformerEncoderModel,
            'TransformerEncoderModel_DAE': TransformerEncoderModel_DAE,
            'CNN_TransformerEncoderModel': CNN_TransformerEncoderModel,
            'FLAT': FLAT
        }
        if hidden_ntag is not None:
            model = models[model_name](config, ntoken, ntag, hidden_ntag, vectors).to(device)
        elif model_name == 'CNN_CRF':
            model = models[model_name](config, ntoken, ntag).to(device)
        elif model_name == 'FLAT':
            model = models[model_name](config, n_bigram, ntoken, ntag, vectors).to(device)
        else:
            model = models[model_name](config, ntoken, ntag, vectors).to(device)
        return model

    def train(self):
        max_f1 = -1
        max_dict = {}
        max_report = {}
        label_report = {}
        loss_list = []
        f1_list = []
        epoch_list = []
        if not os.path.exists('./result/classification_report/{}'.format(self.config.experiment_name)):
            os.mkdir('./result/classification_report/{}'.format(self.config.experiment_name))
            os.mkdir('./result/picture/{}'.format(self.config.experiment_name))
            os.mkdir('./result/data/{}'.format(self.config.experiment_name))
            os.mkdir('./result/data/{}/test_format'.format(self.config.experiment_name))
        logger.info('Loading data ...')
        train_data = self.tool.load_data(self.config.train_path, self.config.is_bioes)
        dev_data = self.tool.load_data(self.config.dev_path, self.config.is_bioes)
        logger.info('Finished load data')
        logger.info('Building vocab ...')
        if self.config.is_pretrained_model:
            with open(self.config.pretrained_vocab, 'r', encoding='utf-8') as vocab_file:
                vocab_list = vocab_file.readlines()
            if self.config.model_name == 'FLAT':
                self.bigram_vocab = self.tool.get_bigram_vocab(train_data, dev_data)
                self.lattice_vocab = self.tool.get_text_vocab(train_data, dev_data)
            else:
                self.word_vocab = self.tool.get_text_vocab(vocab_list)
        else:
            if self.config.model_name == 'FLAT':
                self.bigram_vocab = self.tool.get_bigram_vocab(train_data, dev_data)
                self.lattice_vocab = self.tool.get_text_vocab(train_data, dev_data)
            else:
                self.word_vocab = self.tool.get_text_vocab(train_data, dev_data)
        vectors = self.lattice_vocab.vectors
        self.tag_vocab = self.tool.get_tag_vocab(train_data, dev_data)
        logger.info('Finished build vocab')
        if self.config.is_hidden_tag:
            self.hidden_tag_vocab = self.tool.get_hidden_tag_vocab(train_data, dev_data)
            model = self.init_model(self.config, len(self.word_vocab), len(self.tag_vocab), len(self.hidden_tag_vocab),
                                    vectors=vectors, n_bigram=None)
        elif self.config.model_name == 'FLAT':
            model = self.init_model(self.config, len(self.bigram_vocab), len(self.lattice_vocab), len(self.tag_vocab),
                                    vectors=vectors, n_bigram=None)
        else:
            model = self.init_model(self.config, len(self.word_vocab), len(self.tag_vocab), None, vectors=vectors,
                                    n_bigram=None)
        # model.load_state_dict(torch.load(self.config.model_path.format(self.config.experiment_name)))
        self.model = model
        logger.info('Building iterator ...')
        train_iter = self.tool.get_iterator(train_data, batch_size=self.config.batch_size)
        dev_iter = self.tool.get_iterator(dev_data, batch_size=self.config.batch_size)
        logger.info('Finished build iterator')
        optimizer = optim.Adam(model.parameters(), lr=self.config.learning_rate, weight_decay=1e-5)
        logger.info('Begining train ...')
        for epoch in range(self.config.epoch):
            model.train()
            acc_loss = 0
            for index, iter in enumerate(tqdm(train_iter)):
                if iter.tag.shape[1] == self.config.batch_size:
                    optimizer.zero_grad()
                    if self.config.model_name == 'FLAT':
                        bigiam = iter.bigram[0]
                        lattice = iter.lattice[0]
                        lattice_len = iter.lattice[1]
                        tag = iter.tag
                        loss = model.loss(bigiam, lattice, lattice_len, tag)
                    else:
                        text = iter.text[0]
                        tag = iter.tag
                        text_len = iter.text[1]
                        if self.config.is_hidden_tag:
                            hidden_tag = iter.hidden_tag
                            loss = model.loss(text, text_len, tag, hidden_tag)
                        else:
                            loss = model.loss(text, text_len, tag)
                    acc_loss += loss.view(-1).cpu().data.tolist()[0]
                    loss.backward()
                    optimizer.step()
            f1, report_dict, entity_prf_dict = self.eval(dev_iter)
            loss_list.append(acc_loss)
            f1_list.append(f1)
            epoch_list.append(epoch + 1)
            logger.info('epoch:{}   loss:{}   weighted avg:{}'.format(epoch, acc_loss, report_dict['weighted avg']))
            if f1 > max_f1:
                max_f1 = f1
                label_report = report_dict['weighted avg']
                max_dict = entity_prf_dict['average']
                max_report = entity_prf_dict
                torch.save(model.state_dict(), './save_model/{}.pkl'.format(self.config.experiment_name))
                logger.info(
                    'The best model saved has entity-f1:{}   label-f1:{}'.format(max_f1, label_report['f1-score']))
        logger.info('Finished train')
        logger.info('Max_f1 avg : {}'.format(max_dict))
        # with codecs.open('./result/classification_report/{}/pred_info.txt'.format(config.experiment_name), 'w',
        #                  encoding='utf-8') as f:
        #     f.write(max_dict+ '\n' + label_report)
        self.tool.write_csv(max_report, label_report)
        self.tool.show_1y(epoch_list, loss_list, 'loss')
        self.tool.show_1y(epoch_list, f1_list, 'f1')
        # 稍后处理
        # tool.show_labels_f1_bar_divide(max_report)

    def eval(self, dev_iter):
        self.model.eval()
        tag_true_all = []
        tag_pred_all = []
        entity_prf_dict = {}
        model = self.model
        entities_total = {'origin_place': {'TP': 0, 'S': 0, 'G': 0, 'p': 0, 'r': 0, 'f1': 0},
                          'size': {'TP': 0, 'S': 0, 'G': 0, 'p': 0, 'r': 0, 'f1': 0},
                          'transfered_place': {'TP': 0, 'S': 0, 'G': 0, 'p': 0, 'r': 0, 'f1': 0}}
        with codecs.open('./result/classification_report/{}/pred_info.txt'.format(config.experiment_name), 'w',
                         encoding='utf-8') as f:
            f.write('我要O泡果奶哦哦哦~~~' + '\n')
        for index, iter in enumerate(tqdm(dev_iter)):
            if iter.tag.shape[1] == self.config.batch_size:
                if self.config.model_name == 'FLAT':
                    bigram = iter.bigram[0]
                    lattice = iter.lattice[0]
                    bigram_len = iter.bigram[1]
                    lattice_len = iter.lattice[1]
                    tag = iter.tag.permute(1, 0)
                    result = model(bigram, lattice, lattice_len)
                    for i, result_list in zip(range(bigram.size(1)), result):
                        text1 = lattice.permute(1, 0)
                        sentence = [self.lattice_vocab.itos[w] for w in text1[i][:bigram_len[i]]]
                        tag_list = tag[i][:bigram_len[i]]
                        assert len(tag_list) == len(result_list), 'tag_list: {} != result_list: {}'.format(
                            len(tag_list), len(result_list))
                        tag_true = [self.tag_vocab.itos[k] for k in tag_list]
                        tag_true_all.extend(tag_true)
                        tag_pred = [self.tag_vocab.itos[k] for k in result_list]
                        tag_pred_all.extend(tag_pred)
                        entities = self.tool._evaluate(self.config.is_bioes, tag_true=tag_true, tag_pred=tag_pred,
                                                       sentence=sentence)
                        assert len(entities_total) == len(entities), 'entities_total: {} != entities: {}'.format(
                            len(entities_total), len(entities))
                        for entity in entities_total:
                            entities_total[entity]['TP'] += entities[entity]['TP']
                            entities_total[entity]['S'] += entities[entity]['S']
                            entities_total[entity]['G'] += entities[entity]['G']
                else:
                    text = iter.text[0]
                    tag = torch.transpose(iter.tag, 0, 1)
                    text_len = iter.text[1]
                    result = model(text, text_len)
                    for i, result_list in zip(range(text.size(1)), result):
                        text1 = text.permute(1, 0)
                        sentence = [self.word_vocab.itos[w] for w in text1[i][:text_len[i]]]
                        tag_list = tag[i][:text_len[i]]
                        assert len(tag_list) == len(result_list), 'tag_list: {} != result_list: {}'.format(
                            len(tag_list), len(result_list))
                        tag_true = [self.tag_vocab.itos[k] for k in tag_list]
                        tag_true_all.extend(tag_true)
                        tag_pred = [self.tag_vocab.itos[k] for k in result_list]
                        tag_pred_all.extend(tag_pred)
                        entities = self.tool._evaluate(self.config.is_bioes, tag_true=tag_true, tag_pred=tag_pred,
                                                       sentence=sentence)
                        assert len(entities_total) == len(entities), 'entities_total: {} != entities: {}'.format(
                            len(entities_total), len(entities))
                        for entity in entities_total:
                            entities_total[entity]['TP'] += entities[entity]['TP']
                            entities_total[entity]['S'] += entities[entity]['S']
                            entities_total[entity]['G'] += entities[entity]['G']
        TP = 0
        S = 0
        G = 0
        print('\n--------------------------------------------------')
        print('\tp\t\t\tr\t\t\tf1\t\t\tlabel_type')
        for entity in entities_total:
            entities_total[entity]['p'] = entities_total[entity]['TP'] / entities_total[entity]['S'] \
                if entities_total[entity]['S'] != 0 else 0
            entities_total[entity]['r'] = entities_total[entity]['TP'] / entities_total[entity]['G'] \
                if entities_total[entity]['G'] != 0 else 0
            entities_total[entity]['f1'] = 2 * entities_total[entity]['p'] * entities_total[entity]['r'] / \
                                           (entities_total[entity]['p'] + entities_total[entity]['r']) \
                if entities_total[entity]['p'] + entities_total[entity]['r'] != 0 else 0
            print('\t{:.3f}\t\t{:.3f}\t\t{:.3f}\t\t{}'.format(entities_total[entity]['p'], entities_total[entity]['r'],
                                                              entities_total[entity]['f1'], entity))
            entity_dict = {'precision': entities_total[entity]['p'], 'recall': entities_total[entity]['r'],
                           'f1-score': entities_total[entity]['f1'], 'support': ''}
            entity_prf_dict[entity] = entity_dict
            TP += entities_total[entity]['TP']
            S += entities_total[entity]['S']
            G += entities_total[entity]['G']
        p = TP / S if S != 0 else 0
        r = TP / G if G != 0 else 0
        f1 = 2 * p * r / (p + r) if p + r != 0 else 0
        print('\t{:.3f}\t\t{:.3f}\t\t{:.3f}\t\taverage'.format(p, r, f1))
        print('--------------------------------------------------')
        entity_prf_dict['average'] = {'precision': p, 'recall': r, 'f1-score': f1, 'support': ''}
        labels = []
        for index, label in enumerate(self.tag_vocab.itos):
            labels.append(label)
        labels.remove('O')
        prf_dict = classification_report(tag_true_all, tag_pred_all, labels=labels, output_dict=True)
        return f1, prf_dict, entity_prf_dict

    def predict_sentence(self, model_name=None):
        if model_name is None:
            model_name = self.config.model_path.format(self.config.experiment_name)
        train_data = self.tool.load_data(self.config.train_path, self.config.is_bioes)
        dev_data = self.tool.load_data(self.config.dev_path, self.config.is_bioes)
        logger.info('Finished load data')
        logger.info('Building vocab ...')
        model = None
        # if self.config.is_pretrained_model:
        #     with open(self.config.pretrained_vocab, 'r', encoding='utf-8') as vocab_file:
        #         vocab_list = vocab_file.readlines()
        #     word_vocab = self.tool.get_text_vocab(vocab_list)
        # else:
        #     word_vocab = self.tool.get_text_vocab(train_data, dev_data)
        # vectors = word_vocab.vectors
        # tag_vocab = self.tool.get_tag_vocab(train_data, dev_data)
        # logger.info('Finished build vocab')
        # if self.config.is_hidden_tag:
        #     self.hidden_tag_vocab = self.tool.get_hidden_tag_vocab(train_data, dev_data)
        #     model = self.init_model(self.config, len(self.word_vocab), len(self.tag_vocab), len(self.hidden_tag_vocab),
        #                             vectors=vectors)
        # else:
        #     model = self.init_model(self.config, len(word_vocab), len(tag_vocab), None, vectors=vectors)
        # model.load_state_dict(torch.load(model_name))
        if self.config.is_pretrained_model:
            with open(self.config.pretrained_vocab, 'r', encoding='utf-8') as vocab_file:
                vocab_list = vocab_file.readlines()
            word_vocab = self.tool.get_text_vocab(vocab_list)
        else:
            if self.config.model_name == 'FLAT':
                bigram_vocab = self.tool.get_bigram_vocab(train_data, dev_data)
                lattice_vocab = self.tool.get_text_vocab(train_data, dev_data)
            else:
                word_vocab = self.tool.get_text_vocab(train_data, dev_data)
        vectors = lattice_vocab.vectors
        tag_vocab = self.tool.get_tag_vocab(train_data, dev_data)
        logger.info('Finished build vocab')
        if self.config.is_hidden_tag:
            self.hidden_tag_vocab = self.tool.get_hidden_tag_vocab(train_data, dev_data)
            model = self.init_model(self.config, len(word_vocab), len(tag_vocab), len(self.hidden_tag_vocab),
                                    vectors=vectors)
        elif self.config.model_name == 'FLAT':
            model = self.init_model(self.config, len(bigram_vocab), len(lattice_vocab), len(tag_vocab), vectors=vectors,
                                    n_bigram=None)
        else:
            model = self.init_model(self.config, len(word_vocab), len(tag_vocab), None, vectors=vectors)
        model.load_state_dict(torch.load(model_name))
        f = open(self.config.vocab_path, 'r')
        lines = f.readlines()
        w_list = []
        for line in lines:
            splited = line.strip().split(' ')
            w = splited[0]
            w_list.append(w)
        w_trie = Trie()
        for w in w_list:
            w_trie.insert(w)
        while True:
            print('请输入sentence：')
            sentence = input()
            # texts = self.tool.split_text(sentence)
            # tag_pred = []
            # sentence1 = []
            # for text in texts:
            #     sentence1.extend(text)
            #     text = torch.tensor(numpy.array([word_vocab.stoi[word] for word in text], dtype='int64')).unsqueeze(
            #         1).expand(len(text), self.config.batch_size).to(device)
            #     text_len = torch.tensor(numpy.array([len(text)], dtype='int64')).expand(self.config.batch_size).to(
            #         device)
            #     result = model(text, text_len)[0]
            #     for k in result:
            #         tag_pred.append(tag_vocab.itos[k])
            # sentence1 = ''.join(sentence1)
            # i = 0
            if self.config.model_name =='FLAT':

                texts = self.tool.split_text(sentence)
                tag_pred = []
                sentence1 = []
                for text in texts:
                    sentence1.extend(text)
                    bigram1 = get_bigram(text)
                    bigram = torch.tensor(numpy.array([bigram_vocab.stoi[bi] for bi in bigram1], dtype='int64')).unsqueeze(
                        1).expand(len(bigram1), self.config.batch_size).to(device)
                    lattice1 = list(text) + w_trie.get_lexicon(text)
                    lattice = torch.tensor(
                        numpy.array([lattice_vocab.stoi[word] for word in lattice1], dtype='int64')).unsqueeze(
                        1).expand(len(lattice1), self.config.batch_size).to(device)
                    lattice_len = torch.tensor(numpy.array([len(lattice1)], dtype='int64')).expand(
                        self.config.batch_size).to(
                        device)
                    result = model(bigram, lattice, lattice_len)[0]
                    for k in result:
                        tag_pred.append(tag_vocab.itos[k])
            else:
                texts = self.tool.split_text(sentence)
                tag_pred = []
                for text in texts:
                    sentence1.extend(text)
                    text = torch.tensor(numpy.array([word_vocab.stoi[word] for word in text], dtype='int64')).unsqueeze(
                        1).expand(len(text), self.config.batch_size).to(device)
                    text_len = torch.tensor(numpy.array([len(text)], dtype='int64')).expand(self.config.batch_size).to(device)
                    result = model(text, text_len)[0]
                    for k in result:
                        tag_pred.append(tag_vocab.itos[k])
            sentence1 = ''.join(sentence1)
            i = 0
            origin_places = []
            sizes = []
            transfered_places = []
            while i < len(tag_pred):
                start = 0
                end = 0
                kind = None
                if tag_pred[i] != 'O':
                    start = i
                    kind = tag_pred[i][2:]
                    while i + 1 < len(tag_pred) and tag_pred[i + 1][2:] == kind:
                        i += 1
                    end = i + 1
                    if kind == 'origin_place':
                        origin_places.append(sentence1[start:end])
                    elif kind == 'size':
                        sizes.append(sentence1[start:end])
                    else:
                        transfered_places.append(sentence1[start:end])
                i += 1
            # print(sentence1)
            # print(tag_pred)
            for i in range(len(sentence1)):
                print(sentence1[i], tag_pred[i])
            print(origin_places)
            print(sizes)
            print(transfered_places)

    def predict_test(self, path=None, model_name=None, save_path=None):
        if path is None:
            path = self.config.test_path
            model_name = self.config.model_path.format(self.config.experiment_name)
            save_path = self.config.unformated_val_path.format(self.config.experiment_name)
        train_data = self.tool.load_data(self.config.train_path, self.config.is_bioes)
        dev_data = self.tool.load_data(self.config.dev_path, self.config.is_bioes)
        logger.info('Finished load data')
        logger.info('Building vocab ...')
        model = None
        if self.config.is_pretrained_model:
            with open(self.config.pretrained_vocab, 'r', encoding='utf-8') as vocab_file:
                vocab_list = vocab_file.readlines()
            word_vocab = self.tool.get_text_vocab(vocab_list)
        else:
            if self.config.model_name == 'FLAT':
                bigram_vocab = self.tool.get_bigram_vocab(train_data, dev_data)
                lattice_vocab = self.tool.get_text_vocab(train_data, dev_data)
            else:
                word_vocab = self.tool.get_text_vocab(train_data, dev_data)
        vectors = lattice_vocab.vectors
        tag_vocab = self.tool.get_tag_vocab(train_data, dev_data)
        logger.info('Finished build vocab')
        if self.config.is_hidden_tag:
            self.hidden_tag_vocab = self.tool.get_hidden_tag_vocab(train_data, dev_data)
            model = self.init_model(self.config, len(word_vocab), len(tag_vocab), len(self.hidden_tag_vocab),
                                    vectors=vectors)
        elif self.config.model_name == 'FLAT':
            model = self.init_model(self.config, len(bigram_vocab), len(lattice_vocab), len(tag_vocab), vectors=vectors,
                                    n_bigram=None)
        else:
            model = self.init_model(self.config, len(word_vocab), len(tag_vocab), None, vectors=vectors)
        model.load_state_dict(torch.load(model_name))
        wb = load_workbook(filename=path)
        ws = wb['sheet1']
        max_row = ws.max_row
        f = open(self.config.vocab_path, 'r')
        lines = f.readlines()
        w_list = []
        for line in lines:
            splited = line.strip().split(' ')
            w = splited[0]
            w_list.append(w)
        w_trie = Trie()
        for w in w_list:
            w_trie.insert(w)
        for line_num in tqdm(range(max_row - 1)):
            line_num += 2
            sentence = ws.cell(line_num, 1).value

            # index_size = {}
            # chars = ['.', '*', '×', 'X', 'x', 'c', 'C', 'm', 'M']
            # starts = []
            # ends = []
            # i = 0
            # while i < len(sentence):
            #     if sentence[i] in chars or sentence[i].isdigit():
            #         S_start = i
            #         while i + 1 < len(sentence) and (sentence[i + 1] in chars or sentence[i + 1].isdigit()):
            #             i += 1
            #         if sentence[S_start:i + 1].__contains__('M') or sentence[S_start:i + 1].__contains__('m'):
            #             starts.append(S_start)
            #             ends.append(i)
            #         i += 1
            #     else:
            #         i += 1
            # sentence.replace('$', '')
            # new_sentence = [c for c in sentence]
            # width = 0
            # if len(starts) != 0:
            #     for i in range(len(starts)):
            #         start_i = starts[i] - width
            #         index_size[start_i] = sentence[starts[i]:ends[i] + 1]
            #         for j in range(ends[i] - starts[i]):
            #             del new_sentence[start_i]
            #         new_sentence[start_i] = '$'
            #         width += ends[i] - starts[i]
            #         a = 0
            # sentence = ''.join(new_sentence)
            sentence1 = []
            if self.config.model_name =='FLAT':
                texts = self.tool.split_text(sentence)
                tag_pred = []
                for text in texts:
                    sentence1.extend(text)
                    bigram1 = get_bigram(text)
                    bigram = torch.tensor(numpy.array([bigram_vocab.stoi[bi] for bi in bigram1], dtype='int64')).unsqueeze(
                        1).expand(len(bigram1), self.config.batch_size).to(device)
                    lattice1 = list(text) + w_trie.get_lexicon(text)
                    lattice = torch.tensor(
                        numpy.array([lattice_vocab.stoi[word] for word in lattice1], dtype='int64')).unsqueeze(
                        1).expand(len(lattice1), self.config.batch_size).to(device)
                    lattice_len = torch.tensor(numpy.array([len(lattice1)], dtype='int64')).expand(
                        self.config.batch_size).to(
                        device)
                    result = model(bigram, lattice, lattice_len)[0]
                    for k in result:
                        tag_pred.append(tag_vocab.itos[k])
            else:
                texts = self.tool.split_text(sentence)
                tag_pred = []
                for text in texts:
                    sentence1.extend(text)
                    text = torch.tensor(numpy.array([word_vocab.stoi[word] for word in text], dtype='int64')).unsqueeze(
                        1).expand(len(text), self.config.batch_size).to(device)
                    text_len = torch.tensor(numpy.array([len(text)], dtype='int64')).expand(self.config.batch_size).to(device)
                    result = model(text, text_len)[0]
                    for k in result:
                        tag_pred.append(tag_vocab.itos[k])
            sentence1 = ''.join(sentence1)
            i = 0
            origin_places = []
            sizes = []
            transfered_places = []
            while i < len(tag_pred):
                if self.config.is_bioes:
                    start = end = 0
                    if tag_pred[i][:1] == 'B':
                        kind = tag_pred[i][2:]
                        start = i
                        end = i
                        while end + 1 < len(sentence1) and (
                                tag_pred[end + 1][0] == 'I' or tag_pred[end + 1][0] == 'E') and tag_pred[end + 1][
                                                                                                2:] == kind:
                            end += 1
                        if kind == 'origin_place':
                            origin_places.append(sentence1[start:end + 1])
                        elif kind == 'size':
                            sizes.append(sentence1[start:end + 1])
                        else:
                            transfered_places.append(sentence1[start:end + 1])
                        i = end + 1
                    elif tag_pred[i][:1] == 'E':
                        kind = tag_pred[i][2:]
                        start = i
                        end = i
                        if kind == 'origin_place':
                            origin_places.append(sentence1[start:end + 1])
                        elif kind == 'size':
                            # sizes.append(index_size[start])
                            sizes.append(sentence1[start:end + 1])
                        else:
                            transfered_places.append(sentence1[start:end + 1])
                        i += 1
                    else:
                        i += 1
                else:
                    start = end = 0
                    if tag_pred[i][:1] == 'B':
                        kind = tag_pred[i][2:]
                        start = end = i
                        while end + 1 < len(sentence1) and tag_pred[end + 1][0] == 'I' and tag_pred[end + 1][
                                                                                           2:] == kind:
                            end += 1
                        if kind == 'origin_place':
                            origin_places.append(sentence1[start:end + 1])
                        elif kind == 'size':
                            # sizes.append(index_size[start])
                            sizes.append(sentence1[start:end + 1])
                        else:
                            transfered_places.append(sentence1[start:end + 1])
                        i = end + 1
                    else:
                        i += 1

                # if tag_pred[i]!='O':
                #     start = i
                #     kind = tag_pred[i][2:]
                #     while i+1<len(tag_pred) and tag_pred[i+1][2:]==kind:
                #         i+=1
                #     end = i + 1
                #     if kind == 'origin_place':
                #         origin_places.append(sentence1[start:end])
                #     elif kind == 'size':
                #         sizes.append(index_size[start])
                #     else:
                #         transfered_places.append(sentence1[start:end])
                # i+=1
            for places in [origin_places, sizes, transfered_places]:
                for place in places:
                    if place == []:
                        places.remove(place)
            ws.cell(line_num, 2).value = ','.join(list(set(origin_places)))
            ws.cell(line_num, 3).value = ','.join(list(set(sizes)))
            ws.cell(line_num, 4).value = ','.join(list(set(transfered_places)))
        wb.save(filename=save_path)
        logger.info('Finished Predicting...')

    def test_format_result(self):
        self.train_data = self.tool.load_data(self.config.train_path)
        self.dev_data = self.tool.load_data(self.config.dev_path)
        tag_vocab = self.tool.get_tag_vocab(self.train_data, self.dev_data)
        self.predict_test(path=self.config.dev_path,
                          save_path=self.config.test_unformated_val_path.format(self.config.experiment_name))
        tag_true = []
        tag_formated_pred = []
        tag_unformated_pred = []
        format_result(path=self.config.test_unformated_val_path.format(self.config.experiment_name),
                      save_path=self.config.test_formated_val_path.format(self.config.experiment_name))
        dev_data = self.tool.load_data(self.config.dev_path)
        formated_dev_data = self.tool.load_data(self.config.test_formated_val_path.format(self.config.experiment_name))
        unformated_dev_data = self.tool.load_data(
            self.config.test_unformated_val_path.format(self.config.experiment_name))
        assert len(dev_data.examples) == len(
            unformated_dev_data.examples), 'train_dev_data:{} != unformated_train_dev_data:{}'.format(
            len(dev_data.examples), len(unformated_dev_data.examples))
        assert len(dev_data.examples) == len(
            formated_dev_data.examples), 'train_dev_data:{} != formated_train_dev_data:{}'.format(
            len(dev_data.examples), len(formated_dev_data.examples))
        for example1 in dev_data.examples:
            tag_true.extend(example1.tag)
        for example2 in formated_dev_data.examples:
            tag_formated_pred.extend(example2.tag)
        for example3 in unformated_dev_data.examples:
            tag_unformated_pred.extend(example3.tag)
        # the eval of unformated result
        for i in range(len(dev_data)):
            pass
        assert len(tag_true) == len(tag_unformated_pred), 'tag_true:{} != tag_pred:{}'.format(len(tag_true),
                                                                                              len(tag_unformated_pred))
        assert len(tag_true) == len(tag_formated_pred), 'tag_true:{} != tag_pred:{}'.format(len(tag_true),
                                                                                            len(tag_formated_pred))
        labels = []
        for index, label in enumerate(tag_vocab.itos):
            labels.append(label)
        labels.remove('O')
        prf_dict_formated = classification_report(tag_true, tag_formated_pred, labels=labels, output_dict=True)
        prf_dict_unformated = classification_report(tag_true, tag_unformated_pred, labels=labels, output_dict=True)
        # the eval of formated result
        logger.info('unformated report{}'.format(prf_dict_formated['weighted avg']))
        logger.info('formated report{}'.format(prf_dict_unformated['weighted avg']))


if __name__ == '__main__':
    ee = EE()
    # ee.train()
    # 可单独运行
    ee.predict_test()
    # ee.predict_sentence()
    # 可单独运行
    # ee.test_format_result()

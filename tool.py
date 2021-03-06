# -*- coding: utf-8 -*-
# @Author   : Just-silent
# @time     : 2020/4/26 10:25
import codecs
import os
import re
import json
import torch
import collections
from tqdm import tqdm
import logging
import random
import numpy as np
import pandas as pd
from config import config, device
from transformers import *
from random import shuffle
import matplotlib.pyplot as plt
from torchtext.vocab import Vectors
from openpyxl import load_workbook, Workbook
from torchtext.data import Field, Dataset, Example, BucketIterator

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()

def x_tokenizer(sentence):
    return [word for word in sentence]

def y_tokenizer(tag: str):
    return [tag]


TEXT = Field(sequential=True, use_vocab=True, tokenize=x_tokenizer, include_lengths=True)
TAG = Field(sequential=True, tokenize=y_tokenizer, use_vocab=True, is_target=True, pad_token=None)
Hidden_TAG = Field(sequential=True, tokenize=y_tokenizer, use_vocab=True, is_target=True, pad_token=None)
bigram = Field(sequential=True, use_vocab=True, tokenize=x_tokenizer, include_lengths=True)
lattice = Field(sequential=True, use_vocab=True, tokenize=x_tokenizer, include_lengths=True)
Fields1 = [('text', TEXT), ('tag', TAG), ('hidden_tag', Hidden_TAG)]
Fields2 = [('text', TEXT), ('tag', TAG)]
Fields3 = [('bigram', bigram), ('lattice',lattice), ('tag', TAG)]

def get_all_tag_bioes(sentence, origin_places, sizes, transfered_places):
    len_sentence = len(sentence)
    tag = ['O' for i in range(len_sentence)]
    tag_kinds = ['origin_place', 'size', 'transfered_place']
    for i, columns in enumerate([origin_places, sizes, transfered_places]):
        if columns is not None:
            if i==0 or i==2:
                for column in columns.split(','):
                    starts = tool.find_all_index(sentence, column)
                    ends = [start + len(column) -1 for start in starts]

                    for j in range(len(starts)):
                        x = starts[j]
                        if starts[j] == ends[j]:
                            tag[x] = 'S_{}'.format(tag_kinds[i])
                        else:
                            tag[x] = 'B_{}'.format(tag_kinds[i])
                            x += 1
                            while x < ends[j]:
                                tag[x] = 'I_{}'.format(tag_kinds[i])
                                x += 1
                            while x == ends[j]:
                                tag[x] = 'E_{}'.format(tag_kinds[i])
                                x += 1
            else:
                for column in columns.split(','):
                    start = 0
                    end = 0
                    start = sentence.find(column)
                    end = start + len(column) -1
                    if start==-1:
                        print('未找到')
                    tag[start] = 'B_{}'.format(tag_kinds[i])
                    start += 1
                    while start < end:
                        tag[start] = 'I_{}'.format(tag_kinds[i])
                        start += 1
                    while start == end:
                        tag[start] = 'E_{}'.format(tag_kinds[i])
                        start += 1
    return tag

def get_tag(sentence, origin_places, sizes, transfered_places):
    len_sentence = len(sentence)
    tag = ['O' for i in range(len_sentence)]
    tag_kinds = ['origin_place', 'size', 'transfered_place']
    for i, columns in enumerate([origin_places, sizes, transfered_places]):
        if columns is not None:
            for column in columns.split(','):
                # 如果能够直接找到
                starts = tool.find_all_index(sentence, column)
                ends = [start + len(column) - 1 for start in starts]
                # 不能直接找到
                size_chars = ['C', 'M', 'c', 'm', '*', '×', '.', ' ', 'X']
                if starts == []:
                    # 如果是一维数据（eg：35cm，而不是35cm*35cm）
                    if not column.__contains__('*') and not column.__contains__('×'):
                        num = ''
                        for x in column:
                            if x.isdigit():
                                num+=x
                        start = sentence.find(num)
                        end = start+1
                        while True:
                            if sentence[end] in size_chars[:4] or sentence[end].isdigit():
                                end+=1
                            else:
                                break
                    # 如果是二维数据（eg：不是35cm，而是35cm*35cm）
                    elif column.__contains__('*') or column.__contains__('×'):
                        if column.__contains__('*'):
                            sub1, sub2 = column.split('*')
                        else:
                            sub1, sub2 = column.split('×')
                        num1 = ''
                        for x in sub1:
                            if x.isdigit() or x == '.':
                                num1 += x
                        num2 = ''
                        for x in sub2:
                            if x.isdigit() or x == '.':
                                num2 += x
                        start = sentence.find(num1)
                        start2 = sentence.find(num2)
                        if start2 - start >10:
                            print('距离过大，寻找错误')
                        else:
                            end = start
                            while True:
                                if sentence[end+1] in size_chars or sentence[end+1].isdigit():
                                    end += 1
                                else:
                                    break
                for j in range(len(starts)):
                    start = starts[j]
                    end = ends[j]
                    tag[start] = 'B_{}'.format(tag_kinds[i])
                    start += 1
                    while start<=end:
                        tag[start] = 'I_{}'.format(tag_kinds[i])
                        start+=1
    return tag

def get_all_tag(sentence, origin_places, sizes, transfered_places):
    len_sentence = len(sentence)
    tag = ['O' for i in range(len_sentence)]
    tag_kinds = ['origin_place', 'size', 'transfered_place']
    for i, columns in enumerate([origin_places, sizes, transfered_places]):
        if columns is not None:
            if i==0 or i==2:
                for column in columns.split(','):
                    starts = tool.find_all_index(sentence, column)
                    ends = [start + len(column) -1 for start in starts]
                    for j in range(len(starts)):
                        x = starts[j]
                        tag[x] = 'B_{}'.format(tag_kinds[i])
                        x += 1
                        while x <= ends[j]:
                            tag[x] = 'I_{}'.format(tag_kinds[i])
                            x += 1
            else:
                for column in columns.split(','):
                    start = 0
                    end = 0
                    start = sentence.find(column)
                    end = start + len(column) -1
                    if start==-1:
                        print('未找到')
                    tag[start] = 'B_{}'.format(tag_kinds[i])
                    start += 1
                    while start <= end:
                        tag[start] = 'I_{}'.format(tag_kinds[i])
                        start += 1
    return tag

def get_all_tag_size(sentence, origin_places, sizes, transfered_places):
    # 在原句子查找所有的size格式的位置
    chars = ['.', '*', '×', 'X', 'x', 'c', 'C', 'm', 'M']
    i = 0
    starts = []
    ends = []
    if sizes is not None:
        sizes_s = sizes.split(',')
    kth_sizes = []
    kth = 0
    while i<len(sentence):
        if sentence[i] in chars or sentence[i].isdigit():
            S_start = i
            while i+1<len(sentence) and (sentence[i+1] in chars or sentence[i+1].isdigit()):
                i+=1
            if sentence[S_start:i+1].__contains__('M') or sentence[S_start:i+1].__contains__('m'):
                starts.append(S_start)
                ends.append(i)
                if sizes is not None and sentence[S_start:i+1] in sizes_s:
                    kth_sizes.append(kth)
                kth+=1
            i+=1
        else:
            i+=1
    sentence.replace('$','')
    new_sentence = [c for c in sentence]
    width = 0
    if len(starts)!=0:
        for i in range(len(starts)):
            start_i = starts[i]-width
            for j in range(ends[i]-starts[i]):
                del new_sentence[start_i]
            new_sentence[start_i] = '$'
            width+=ends[i]-starts[i]
            a=0
    sentence = ''.join(new_sentence)
    # 用$替换原size，获得新句子，并找到标注的原位置    len_sentence = len(sentence)
    tag = ['O' for i in range(len(sentence))]
    index_s = tool.find_all_index(sentence,'$')
    if len(kth_sizes) != 0:
        for i in range(len(kth_sizes)):
            tag[index_s[kth_sizes[i]]] = 'E_size'
    tag_kinds = ['origin_place', 'size', 'transfered_place']
    for i, columns in enumerate([origin_places, sizes, transfered_places]):
        if columns is not None:
            if i==0 or i==2:
                for column in columns.split(','):
                    starts = tool.find_all_index(sentence, column)
                    ends = [start + len(column) -1 for start in starts]
                    for j in range(len(starts)):
                        x = starts[j]
                        tag[x] = 'B_{}'.format(tag_kinds[i])
                        x += 1
                        while x < ends[j]:
                            tag[x] = 'I_{}'.format(tag_kinds[i])
                            x += 1
                        tag[ends[j]] = 'E_{}'.format(tag_kinds[i])
    return tag, new_sentence

def get_bigram(sentence):
    bigram = []
    for i in range(len(sentence)-1):
        bigram.append(sentence[i:i+2])
    bigram.append(sentence[-1]+'end')
    return bigram

def get_flat(sentence, origin_places, sizes, transfered_places, w_trie):
    # 1.bigram
    bigram = get_bigram(sentence)
    # 2.lattice
    lattice = list(sentence) + w_trie.get_lexicon(sentence)
    # 3.tag
    tag = get_all_tag(sentence, origin_places, sizes, transfered_places)
    return bigram, lattice, tag

class EEDataset(Dataset):
    def __init__(self, path, is_bioes, fields, encoding="utf-8", **kwargs):
        examples = []
        wb = load_workbook(filename=path)
        ws = wb['sheet1']
        max_row = ws.max_row
        if config.model_name == 'FLAT':
            f = open(config.vocab_path, 'r')
            lines = f.readlines()
            w_list = []
            for line in lines:
                splited = line.strip().split(' ')
                w = splited[0]
                w_list.append(w)
            w_trie = Trie()
            for w in w_list:
                w_trie.insert(w)
        for line_num in tqdm(range(max_row-1)):
            line_num = line_num+2
            sentence, origin_places ,sizes ,transfered_places = ws.cell(line_num, 1).value, ws.cell(line_num, 2).value, ws.cell(line_num, 3).value, ws.cell(line_num, 4).value
            hidden_tag = [0]
            if not (origin_places is None and sizes is None and transfered_places is None):
                hidden_tag = [1]
            if sentence is not None:
                if is_bioes:
                    if config.model_name == 'FLAT':
                               # * .
                        bigram, lattice, tag = get_flat(sentence, origin_places, sizes, transfered_places, w_trie)
                    # size占位符 bioes
                    # tag_list, sentence_list = get_all_tag_size(sentence, origin_places, sizes, transfered_places)
                    # size非占位符 bioes
                    # tag_list = get_all_tag_bioes(sentence, origin_places, sizes, transfered_places)
                    # size非占位符 bio
                    tag_list = get_all_tag(sentence, origin_places, sizes, transfered_places)
                    # tag_list = get_tag(sentence, origin_places, sizes, transfered_places)
                    sentence_list = [x for x in sentence]
                else:
                    if config.model_name == 'FLAT':

                        bigram, lattice, tag = get_flat(sentence, origin_places, sizes, transfered_places, w_trie)
                    # size占位符 bioes
                    # tag_list, sentence_list = get_all_tag_size(sentence, origin_places, sizes, transfered_places)
                    # size非占位符 bioes
                    # tag_list = get_all_tag_bioes(sentence, origin_places, sizes, transfered_places)
                    # size非占位符 bio
                    tag_list = get_all_tag(sentence, origin_places, sizes, transfered_places)
                    # tag_list = get_tag(sentence, origin_places, sizes, transfered_places)
                    sentence_list = [x for x in sentence]
                if config.model_name == 'BiLSTM_CRF_hidden_tag':
                    examples.append(Example.fromlist((sentence_list, tag_list, hidden_tag), fields))
                elif config.model_name == 'FLAT':
                    examples.append(Example.fromlist((bigram, lattice, tag_list), fields))
                else:
                    examples.append(Example.fromlist((sentence_list, tag_list), fields))

        super(EEDataset, self).__init__(examples, fields, **kwargs)

class TrieNode:
    def __init__(self):
        self.children = collections.defaultdict(TrieNode)
        self.is_w = False

class Trie:
    def __init__(self):
        self.root = TrieNode()

    def insert(self,w):
        current = self.root
        for c in w:
            current = current.children[c]
        current.is_w = True

    def search(self,w):
        current = self.root
        for c in w:
            current = current.children.get(c)
            if current is None:
                return -1
        if current.is_w:
            return 1
        else:
            return 0

    def get_lexicon(self,sentence):
        result = []
        for i in range(len(sentence)):
            current = self.root
            for j in range(i, len(sentence)):
                current = current.children.get(sentence[j])
                if current is None:
                    break

                if current.is_w:
                    result.append(sentence[i:j+1])

        return result

class Tool():
    def __init__(self, config=None):
        if config is not None:
            if config.is_hidden_tag:
                self.Fields = Fields1
            elif config.model_name == 'FLAT':
                self.Fields = Fields3
            else:
                self.Fields = Fields2

    def load_data(self, path: str, is_bioes):
        fields = self.Fields
        dataset = EEDataset(path, is_bioes, fields=fields)
        return dataset

    def get_text_vocab(self, *dataset):
        if config.model_name == 'FLAT':
            lattice.build_vocab(*dataset)
            return lattice.vocab
        if config.is_vector is False:
            TEXT.build_vocab(*dataset)
        else:
            vec = Vectors(name=config.vector)
            TEXT.build_vocab(*dataset,
                 max_size=3000,
                 min_freq=1,
                 vectors=vec,  #vects替换为None则不使用词向量
                 unk_init = torch.Tensor.normal_)
        return TEXT.vocab

    def get_tag_vocab(self, *dataset):
        TAG.build_vocab(*dataset)
        return TAG.vocab

    def get_bigram_vocab(self, *dataset):
        bigram.build_vocab(*dataset)
        return bigram.vocab

    def get_hidden_tag_vocab(self, *dataset):
        Hidden_TAG.build_vocab(*dataset)
        return Hidden_TAG.vocab

    def get_iterator(self, dataset: Dataset, batch_size=1,
                     sort_key=lambda x: len(x.text), sort_within_batch=True):
        iterator = BucketIterator(dataset, batch_size=batch_size, sort_key=sort_key,
                              sort_within_batch=sort_within_batch, device=device)
        return iterator

    def _evaluate(self, is_bioes, tag_true, tag_pred, sentence):
        """
        先对true进行还原成 [{}] 再对pred进行还原成 [{}]
        :param tag_true: list[]
        :param tag_pred: list[]
        :return:
        """
        true_list = self._build_list_dict(is_bioes, _len=len(tag_true), _list=tag_true, sentence=sentence)
        pred_list = self._build_list_dict(is_bioes, _len=len(tag_pred), _list=tag_pred, sentence=sentence)
        entities = {'origin_place': {'TP': 0, 'S': 0, 'G': 0},
                    'size': {'TP': 0, 'S': 0, 'G': 0},
                    'transfered_place': {'TP': 0, 'S': 0, 'G': 0}}
        for true in true_list:
            label_type = true['label_type']
            entities[label_type]['G'] += 1
        for pred in pred_list:
            label_type = pred['label_type']
            label_name = pred['name']
            label_start = pred['start_pos']
            label_end = pred['end_pos']
            entities[label_type]['S'] += 1
            for true in true_list:
                if label_type == true['label_type'] and label_name == true['name'] and label_start == true['start_pos'] and label_end == true['end_pos']:
                    entities[label_type]['TP'] += 1
        self.record_pred_info(sentence=sentence, true_list=true_list, pred_list=pred_list,path = './result/classification_report/{}/pred_info.txt'.format(config.experiment_name))
        return entities

    def record_pred_info(self, sentence=None, true_list=None, pred_list=None, path=None):
        pred_false = []
        un_pred = []
        sentence = ''.join(sentence)
        for pred in pred_list:
            start_pos = pred['start_pos']
            end_pos = pred['end_pos']
            label_type = pred['label_type']
            _bool = False
            for true in true_list:
                if label_type == true['label_type'] and start_pos == true['start_pos'] and end_pos == true['end_pos']:
                    _bool = True
                    break
            if not _bool:
                pred_false.append({'entity': sentence[start_pos: end_pos+1], 'label_type': label_type})
        for true in true_list:
            start_pos = true['start_pos']
            end_pos = true['end_pos']
            label_type = true['label_type']
            _bool = False
            for pred in pred_list:
                if label_type == pred['label_type'] and start_pos == pred['start_pos'] and end_pos == pred['end_pos']:
                    _bool = True
                    break
            if not _bool:
                un_pred.append({'entity': sentence[start_pos: end_pos+1], 'label_type': label_type})
        pred_dict = {'sentence': sentence, 'pred_false': pred_false, 'un_pred': un_pred}

        # for pred in pred_list:
        #
        #     _bool = False
        #     for true in true_list:
        #         if true == pred:
        #             _bool = True
        #             break
        #         if not _bool and pred != '':
        #             pred_false.append({'true': true, 'pred':pred })
        # for true in true_list:
        #     _bool = False
        #     for pred in pred_list:
        #         if pred == true:
        #             _bool = True
        #             break
        #         if not _bool and pred == '':
        #             un_pred.append({'true': true})
        # pred_dict = {'sentence': sentence, 'pred_false': pred_false, 'un_pred': un_pred}
        with codecs.open(path, 'a', encoding='utf-8') as f:
            f.write(json.dumps(pred_dict, ensure_ascii=False) + '\n')

    def _build_list_dict(self, is_bioes, _len, _list, sentence):
        build_list = []
        tag_dict = {'origin_place': 'origin_place',
                    'size': 'size',
                    'transfered_place': 'transfered_place'}
        i = 0
        if is_bioes:
            while i < _len:
                if _list[i][0] == 'B':
                    label_type = _list[i][2:]
                    start_pos = i
                    end_pos = start_pos
                    while end_pos + 1 < _len and (_list[end_pos + 1][0] == 'I' or _list[end_pos + 1][0] == 'E') and _list[end_pos + 1][
                                                           2:] == label_type:
                        end_pos += 1
                    build_list.append(
                        {'name': ''.join(sentence[start_pos:end_pos + 1]), 'start_pos':start_pos, 'end_pos':end_pos, 'label_type': tag_dict[label_type]})
                    i = end_pos + 1
                elif _list[i][0] == 'E':
                    build_list.append(
                        {'name': ''.join(sentence[i:i + 1]), 'start_pos':i, 'end_pos':i, 'label_type': _list[i][2:]})
                    i+=1
                else:
                    i+=1
        else:
            while i<_len:
                if _list[i][0] == 'B':
                    label_type = _list[i][2:]
                    start_pos = i
                    end_pos = start_pos
                    if end_pos+1<_len and _list[end_pos+1][0] != 'I':
                        end_pos+=1
                    else:
                        while end_pos+1 < _len and _list[end_pos+1][0] == 'I' and _list[end_pos+1][2:] == label_type:
                            end_pos += 1
                    build_list.append({'name': ''.join(sentence[start_pos:end_pos+1]), 'start_pos':start_pos, 'end_pos':end_pos, 'label_type': tag_dict[label_type]})
                    i = end_pos+1
                else:
                    i+=1
        result = []
        for dict1 in build_list:
            if dict1 not in result:
                result.append(dict1)
        return result

    def show_1y(self, list_x, list_y, name):
        fig, ax = plt.subplots()
        plt.xlabel('The updating of {}'.format(name))
        plt.ylabel(name)
        plt.plot(list_x, list_y, label=name)

        """set interval for y label"""
        # yticks = range(0, 100, 10)
        # ax.set_yticks(yticks)

        """set min and max value for axes"""
        # ax.set_ylim([0, 100])
        # ax.set_xlim([1, len(list_x)])
        plt.grid(True)
        plt.legend(bbox_to_anchor=(1.0, 1), loc=1, borderaxespad=0.)
        plt.savefig('./result/picture/{}/{}.jpg'.format(config.experiment_name, name))
        # 展示图片需要点击才能下一步
        # plt.show()

    def write_csv(self, dict, label_dict):
        # 这里补充相应的配置信息 1.model 2.细节信息 3.epoch
        tag_list = []
        p_list = []
        r_list = []
        f1_list = []
        s_list = []
        for i, name in enumerate(dict):
            tag_list.append(name)
            p_list.append(dict[name]['precision'])
            r_list.append(dict[name]['recall'])
            f1_list.append(dict[name]['f1-score'])
            s_list.append(dict[name]['support'])
        tag_list.append('label_averag')
        p_list.append(label_dict['precision'])
        r_list.append(label_dict['recall'])
        f1_list.append(label_dict['f1-score'])
        s_list.append(label_dict['support'])
        dataframe = pd.DataFrame({'name': tag_list, 'precision': p_list, 'recall': r_list, 'f1': f1_list, 's_persent': s_list})
        # dataframe.to_csv('./result/classification_report/{}/report.csv'.format(config.experiment_name), index=False, sep=str(','))
        dataframe.to_excel('./result/classification_report/{}/report.xlsx'.format(config.experiment_name), sheet_name='sheet1')

    def labels_f1_bar(self, x_name, support, f1, pict_name):
    # def labels_f1_bar(self, x_name, support, f1):
        # 这两行代码解决 plt 中文显示的问题
        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.rcParams['axes.unicode_minus'] = False
        bar_width = 5  # 条形宽度
        index_support = np.arange(len(x_name)) * 10  # support条形图的横坐标
        index_f1 = index_support + bar_width  # f1条形图的横坐标
        # 使用两次 bar 函数画出两组条形图
        rects1 = plt.bar(index_support, height=support, width=bar_width, color='b', label='占比')
        rects2 = plt.bar(index_f1, height=f1, width=bar_width, color='g', label='f1')
        # 显示对应柱状图的数值
        for rect in rects1:
            height = rect.get_height()
            plt.text(rect.get_x() + rect.get_width() / 2, height, str(height) + '%', ha='center', va='bottom')
        for rect in rects2:
            height = rect.get_height()
            plt.text(rect.get_x() + rect.get_width() / 2, height, str(height) + '%', ha='center', va='bottom')
        plt.legend()  # 显示图例
        plt.xticks(index_support + bar_width / 2, x_name, rotation=-90)  # 让横坐标轴刻度显示 x_name，index_support + bar_width/2 为横坐标轴刻度的位置
        plt.ylim(0, 1)
        plt.ylabel('present')  # 纵坐标轴标题
        plt.title('各个标签占比及其f1')  # 图形标题
        plt.savefig('./result/picture/{}/{}.jpg'.format(config.experiment_name, pict_name))
        plt.show()

    def find_all_index(self, str1, str2):
        # 子串str2， in str1查找目标字符串
        starts = []
        start = 0
        over = 0
        while True:
            index = str1.find(str2, start, len(str1))
            if index != -1:
                starts.append(index+over)
                if index+len(str2)<=len(str1)-1:
                    str1 = str1[index+len(str2):]
                    over += index+len(str2)
                else:
                    break
            else:
                break
        return starts

    def split_text(self, sentence):
        result1 = []
        result2 = []
        texts = re.split('。', sentence)
        for i in range(len(texts)):
            if texts[i] != '':
                result1.append(texts[i]+'。')
        # for text in result1:
        #     if text.__contains__(';'):
        #         texts = re.split(';', text)
        #         for i in range(len(texts)-1):
        #             if texts[i] != '':
        #                 result2.append(texts[i]+';')
        #         result2.append(texts[len(texts)-1])
        #     elif text.__contains__('；'):
        #         texts = re.split('；', text)
        #         for i in range(len(texts)-1):
        #             if texts[i] != '':
        #                 result2.append(texts[i] + '；')
        #         result2.append(texts[len(texts)-1])
        #     else:
        #         result2.append(text)
        return result1

    def split_describe_conclusion(self, str1):
        i = 1
        while True:
            if re.search('({}\.|{}\、)\D'.format(i, i), str1) is not None:
                i+=1
            else:
                i-=1
                break
        if i==0:
            return str1, None
        str2 = ''
        for j in range(i):
            if j<i-1:
                str2 = str2 + '({}\.|{}\、)\D.*?'.format(j+1, j+1)
            else:
                str2 = str2 + '({}\.|{}\、)\D.*?{}'.format(j+1, j+1, '。')
        result = re.search(str2, str1)
        if result is None:
            return str1, None
        else:
            return str1.replace(result.group(),''), result.group()

tool = Tool()
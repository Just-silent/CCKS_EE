# 优化结果
import numpy as np
from openpyxl import load_workbook
from config import config

# 问题
# 1. 对于部位补全问题，可能存在多个对应的（可能性不大）
# 2. size问题    对于这两个问题，可以采用新的标注方式：对三类名词标注成一个B
def format_result(path='./data/{}/unformated_val.xlsx'.format(config.experiment_name), save_path='./data/{}/formated_val.xlsx'.format(config.experiment_name)):
    with open('../vocab/task2_vocab.val.txt', 'r', encoding='utf-8') as f:
        names = f.readlines()
    # 此处路径需要修改
    wb = load_workbook(path)
    ws = wb['sheet1']
    max_row = ws.max_row
    num = 0
    for i in range(max_row-1):
        line = i+2
        if ws.cell(line, 3).value is not None:
            list_cell3 = ws.cell(line, 3).value.split(',')
            for j in range(len(list_cell3)):
                if list_cell3[j].__contains__('×') or list_cell3[j].__contains__('X') or list_cell3[j].__contains__('*'):
                    unit = ''
                    nums = ''
                    num_strs = list_cell3[j].strip()
                    flag = False
                    for i in range(len(num_strs)):
                        if num_strs[i].isdigit() or num_strs[i] == '.':
                            flag = True
                            nums+=num_strs[i]
                        else:
                            if flag:
                                nums+=','
                                flag = False
                    if 'CM' in num_strs or 'cm' in num_strs:
                        unit = 'CM'
                    else:
                        unit = 'MM'
                    nums = nums[:-1].split(',')
                    result = ''
                    for num_ in nums:
                        result = result + num_ + unit + '×'
                    list_cell3[j] = result[:-1]
                else:
                    unit = ''
                    num_ = ''
                    num_str = list_cell3[j].strip()
                    for c in num_str:
                        if c.isdigit() or c == '.':
                            num_+=c
                    if 'CM' in num_str or 'cm' in num_str:
                        unit = 'CM'
                    else:
                        unit = 'MM'
                    list_cell3[j] = num_+unit
            x = ','.join(list_cell3)
            ws.cell(line, 3).value = ','.join(list_cell3)
        # for j in range(len(names)):
        #     if ws.cell(line, 2).value is not None:
        #         list_cell2 = ws.cell(line, 2).value.split(',')
        #         for k in range(len(list_cell2)):
        #             if names[j].__contains__(list_cell2[k]) & ws.cell(line, 1).value.__contains__(names[j][:-1]):
        #                 list_cell2[k] = names[j][:-1]
        #                 num += 1
        #         ws.cell(line, 2).value = ','.join(list_cell2)
        #     if ws.cell(line, 4).value is not None:
        #         list_cell4 = ws.cell(line, 4).value.split(',')
        #         for k in range(len(list_cell4)):
        #             if names[j].__contains__(list_cell4[k]) & ws.cell(line, 1).value.__contains__(names[j][:-1]):
        #                 list_cell4[k] = names[j][:-1]
        #                 num += 1
        #         ws.cell(line, 4).value = ','.join(list_cell4)
    wb.save(save_path)
    print('优化完成，优化{}次'.format(num))

def eval(path='./data/{}/unformated_val.xlsx'.format(config.experiment_name)):
    wb = load_workbook(path)
    ws = wb['sheet1']
    max_row = ws.max_row
    r = np.array(np.zeros((400,3)))
    for i in range(max_row-1):
        line = i+2
        for i in range(3):
            colum1 = i+2
            colum2 = i+5
            trues = []
            preds = []
            if ws.cell(line, colum1).value is not None and ws.cell(line, colum2).value is not None:
                trues = ws.cell(line, colum1).value.split(',')
                preds = ws.cell(line, colum2).value.split(',')



if __name__ == '__main__':
    format_result()
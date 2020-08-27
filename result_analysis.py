# -*- coding: utf-8 -*-
# @Author   : Just-silent
# @time     : 2020/6/24 21:34

import numpy as np
import torch
from tqdm import tqdm
from tool import tool, logger
from config import config, device
from openpyxl import load_workbook, Workbook
from model import BiLSTM_CRF, BiLSTM_CRF_ATT, CNN_CRF, TransformerEncoderModel

def result_four(name):
    wb = load_workbook('./result/data/{}/analysis.xlsx'.format(name))
    ws = wb['sheet1']
    max_row = ws.max_row
    four = np.array(np.zeros((3, 19, 5)))
    four2 = np.array(np.zeros((3, 19, 5)))
    four_len = np.array(np.zeros((3, 19)))
    for i in range(max_row - 1):
        four1 = np.array(np.zeros((3, 19, 5)))
        line = i + 2
        for j in range(3):
            if ws.cell(line, j + 2).value is None:
                four_len[j, 0] += 1
            if ws.cell(line, j + 2).value is not None:
                tag_len = len(ws.cell(line, j + 2).value.split(','))
                true_tags = ws.cell(line, j + 2).value.split(',')
                four_len[j,tag_len] += 1
                if ws.cell(line, j + 5).value is not None:
                    pred_tags = ws.cell(line, j + 5).value.split(',')
                else:
                    pred_tags = [' ']
                nt = 0 #正确 基于true_tags
                nc = 0 #不全
                nn = 0 #未抽
                nm = 0 #错误 基于true_tags 错误=不全+未抽
                pm = 0 #结果错误
                for true_tag in true_tags:
                    if true_tag not in pred_tags:
                        flag = False
                        nn+=1
                        for pred_tag in pred_tags:
                            if true_tag.__contains__(pred_tag):
                                flag = True
                        if flag:
                            nc+=1
                        else:
                            nm+=1
                    else:
                        nt+=1
                for pred_tag in pred_tags:
                    if pred_tag not in true_tags:
                        if pred_tag != ' ':
                            flag = True
                            for true_tag in true_tags:
                                if true_tag.__contains__(pred_tag):
                                    flag = False
                            if flag:
                                pm+=1
                four1[j][tag_len][0] = nt
                four1[j][tag_len][1] = nc
                four1[j][tag_len][2] = nm
                four1[j][tag_len][3] = nn
                four1[j][tag_len][4] = pm
            elif ws.cell(line, j + 5).value is not None:
                four1[j][0][4] += len(ws.cell(line, j + 5).value.split(','))
        four += four1
    for i in range(3):
        for j in range(19):
            for k in range(5):
                if four_len[i][j]!=0.:
                    four2[i][j][k] = four[i][j][k]/four_len[i][j]
    wb = Workbook()
    wb.remove(wb['Sheet'])
    ws = wb.create_sheet('sheet1')
    ws.cell(1,1,)
    names = ['真实标签个数','正确avg','不全avg','未抽avg','错误avg','预测错误avg']
    line_index = 2
    for i in range(6):
        ws.cell(1, i+1, names[i])
    for i in range(3):
        for j in range(19):
            ws.cell(line_index, 1, j)
            for k in range(5):
                ws.cell(line_index, k+2, four2[i][j][k])
            line_index+=1
        line_index+=1
        for i in range(6):
            ws.cell(line_index, i + 1, names[i])
        line_index+=1
    wb.save('./result/data/{}/analysis_result.xlsx'.format(name))

def replace_char(path):
    wb = load_workbook(path)
    ws = wb.get_sheet_by_name('sheet1')
    max_row = ws.max_row
    for i in range(max_row - 1):
        line = i + 2
        if ws.cell(line, 2).value is not None:
            ws.cell(line, 2).value = ws.cell(line, 2).value.replace('、',',')
        if ws.cell(line, 4).value is not None:
            ws.cell(line, 4).value = ws.cell(line, 4).value.replace('、',',')
    wb.save(path)

if __name__ == '__main__':
    result_four(name=config.experiment_name)
    # replace_char(path='./data/sub_dev.xlsx')
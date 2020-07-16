# -*- coding: utf-8 -*-
# @Author   : Just-silent
# @time     : 2020/6/24 21:34

import numpy as np
import torch
from tqdm import tqdm
from tool import tool, logger
from config import config, device
from openpyxl import load_workbook, Workbook
from model import BiLSTM_CRF, BiLSTM_CRF_ATT, CNN_CRF, TransformerEncoderModel

def write_val_true_pred():
    model_name = config.model_path.format(config.experiment_name)
    save_path = config.analysis_path.format(config.experiment_name)
    train_data = tool.load_data(config.train_path)
    dev_data = tool.load_data(config.dev_path)
    logger.info('Finished load data')
    logger.info('Building vocab ...')
    if config.is_pretrained_model:
        with open(config.pretrained_vocab, 'r', encoding='utf-8') as vocab_file:
            vocab_list = vocab_file.readlines()
        word_vocab = tool.get_text_vocab(vocab_list)
    else:
        word_vocab = tool.get_text_vocab(train_data, dev_data)
    vectors = word_vocab.vectors
    tag_vocab = tool.get_tag_vocab(train_data, dev_data)
    if config.model_name == 'BiLSTM_CRF':
        model = BiLSTM_CRF(config, ntoken=len(word_vocab), ntag=len(tag_vocab), vectors=vectors).to(
            device)
    elif config.model_name == 'TransformerEncoderModel':
        model = TransformerEncoderModel(config, ntoken=len(word_vocab), ntag=len(tag_vocab), vectors=vectors).to(device)
    elif config.model_name == 'CNN_CRF':
        model = CNN_CRF(config, ntoken=len(word_vocab), ntag=len(tag_vocab)).to(device)
    elif config.model_name == 'BiLSTM_CRF_ATT':
        model = BiLSTM_CRF_ATT(config, ntoken=len(word_vocab), ntag=len(tag_vocab), vectors=vectors).to(
            device)
    model.load_state_dict(torch.load(model_name))
    # 需要新建xlsx 七列
    wb_analysis = Workbook()
    analysis_sheet = wb_analysis.create_sheet('sheet1')
    wb_analysis.remove(wb_analysis['Sheet'])
    names = ['原文','原发部位','病灶大小','转移部位','pred_原发部位','pred_病灶大小','pred_转移部位']
    for i in range(len(names)):
        analysis_sheet.cell(1,i+1,names[i])
    wb = load_workbook(filename=config.dev_path)
    ws = wb['sheet1']
    max_row = ws.max_row
    for line_num in tqdm(range(max_row-1)):
        line_num+=2
        sentence = ws.cell(line_num,1).value
        text = torch.tensor(np.array([word_vocab.stoi[word] for word in sentence], dtype='int64')).unsqueeze(1).expand(len(sentence),8).to(device)
        text_len = torch.tensor(np.array([text.size(0)], dtype='int64')).expand(8).to(device)
        result = model(text, text_len)[0]
        tag_pred = []
        for result_list in [result]:
            list1 = [tag_vocab.itos[k] for k in result_list]
            tag_pred.extend(list1)
        i = 0
        origin_places = []
        sizes = []
        transfered_places = []
        # 此处可能存在问题
        while i < len(tag_pred):
            start = 0
            end = 0
            kind = None
            if tag_pred[i]!='O':
                start = i
                kind = tag_pred[i][2:]
                while i+1<len(tag_pred) and tag_pred[i+1][2:]==kind:
                    i+=1
                end = i + 1
                if kind == 'origin_place':
                    origin_places.append(sentence[start:end])
                elif kind == 'size':
                    sizes.append(sentence[start:end])
                else:
                    transfered_places.append(sentence[start:end])
            i+=1
        analysis_sheet.cell(line_num, 1, ws.cell(line_num,1).value)
        analysis_sheet.cell(line_num, 2, ws.cell(line_num,2).value)
        analysis_sheet.cell(line_num, 3, ws.cell(line_num,3).value)
        analysis_sheet.cell(line_num, 4, ws.cell(line_num,4).value)
        analysis_sheet.cell(line_num, 5, ','.join(origin_places))
        analysis_sheet.cell(line_num, 6, ','.join(sizes))
        analysis_sheet.cell(line_num, 7, ','.join(transfered_places))
    wb_analysis.save(filename=save_path)
    logger.info('Finished Predicting...')

def result_four(name):
    wb = load_workbook('./result/data/{}/analysis.xlsx'.format(name))
    ws = wb['sheet1']
    max_row = ws.max_row
    four = np.array(np.zeros((3, 19, 5)))
    four2 = np.array(np.zeros((3, 19, 5)))
    four_len = np.array(np.zeros((3, 19)))
    for i in range(max_row - 1):
        four1 = np.array(np.zeros((3, 19, 5)))
        line = i + 2
        for j in range(3):
            if ws.cell(line, j + 2).value is None:
                four_len[j, 0] += 1
            if ws.cell(line, j + 2).value is not None:
                tag_len = len(ws.cell(line, j + 2).value.split(','))
                true_tags = ws.cell(line, j + 2).value.split(',')
                four_len[j,tag_len] += 1
                if ws.cell(line, j + 5).value is not None:
                    pred_tags = ws.cell(line, j + 5).value.split(',')
                else:
                    pred_tags = [' ']
                nt = 0 #正确 基于true_tags
                nc = 0 #不全
                nn = 0 #未抽
                nm = 0 #错误 基于true_tags 错误=不全+未抽
                pm = 0 #结果错误
                for true_tag in true_tags:
                    if true_tag not in pred_tags:
                        flag = False
                        nn+=1
                        for pred_tag in pred_tags:
                            if true_tag.__contains__(pred_tag):
                                flag = True
                        if flag:
                            nc+=1
                        else:
                            nm+=1
                    else:
                        nt+=1
                for pred_tag in pred_tags:
                    if pred_tag not in true_tags:
                        if pred_tag != ' ':
                            flag = True
                            for true_tag in true_tags:
                                if true_tag.__contains__(pred_tag):
                                    flag = False
                            if flag:
                                pm+=1
                four1[j][tag_len][0] = nt
                four1[j][tag_len][1] = nc
                four1[j][tag_len][2] = nm
                four1[j][tag_len][3] = nn
                four1[j][tag_len][4] = pm
            elif ws.cell(line, j + 5).value is not None:
                four1[j][0][4] += len(ws.cell(line, j + 5).value.split(','))
        four += four1
    for i in range(3):
        for j in range(19):
            for k in range(5):
                if four_len[i][j]!=0.:
                    four2[i][j][k] = four[i][j][k]/four_len[i][j]
    wb = Workbook()
    wb.remove(wb['Sheet'])
    ws = wb.create_sheet('sheet1')
    ws.cell(1,1,)
    names = ['真实标签个数','正确avg','不全avg','未抽avg','错误avg','预测错误avg']
    line_index = 2
    for i in range(6):
        ws.cell(1, i+1, names[i])
    for i in range(3):
        for j in range(19):
            ws.cell(line_index, 1, j)
            for k in range(5):
                ws.cell(line_index, k+2, four2[i][j][k])
            line_index+=1
        line_index+=1
        for i in range(6):
            ws.cell(line_index, i + 1, names[i])
        line_index+=1
    wb.save('./result/data/{}/analysis_result.xlsx'.format(name))


def replace_char(path):
    wb = load_workbook(path)
    ws = wb.get_sheet_by_name('sheet1')
    max_row = ws.max_row
    for i in range(max_row - 1):
        line = i + 2
        if ws.cell(line, 2).value is not None:
            ws.cell(line, 2).value = ws.cell(line, 2).value.replace('、',',')
        if ws.cell(line, 4).value is not None:
            ws.cell(line, 4).value = ws.cell(line, 4).value.replace('、',',')
    wb.save(path)

if __name__ == '__main__':
    write_val_true_pred()
    result_four(name=config.experiment_name)
    # replace_char(path='./data/sub_dev.xlsx')